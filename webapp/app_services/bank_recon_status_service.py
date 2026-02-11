"""
Bank Reconciliation Status Service

Dashboard showing bank account reconciliation status including
unreconciled transactions.
"""

import logging
from io import BytesIO
from typing import Any

import requests

from webapp.time_utils import utcnow_iso

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"


def generate_bank_recon_status(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> dict[str, Any]:
    """
    Generate bank reconciliation status report.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        as_at_date: Date for status check (YYYY-MM-DD)

    Returns:
        Dict with bank accounts, unreconciled items, and status
    """
    try:
        # Fetch bank accounts
        bank_accounts = _fetch_bank_accounts(access_token, tenant_id)

        # Fetch bank summary report
        bank_summary = _fetch_bank_summary(access_token, tenant_id, as_at_date)

        # Fetch unreconciled transactions for each account
        for account in bank_accounts:
            account_id = account.get("account_id")
            unreconciled = _fetch_unreconciled_transactions(
                access_token, tenant_id, account_id
            )
            account["unreconciled_items"] = unreconciled
            account["unreconciled_count"] = len(unreconciled)
            account["unreconciled_amount"] = sum(
                abs(t.get("amount", 0)) for t in unreconciled
            )

            # Update balance from bank summary if available
            summary_balance = bank_summary.get(account.get("code"))
            if summary_balance is not None:
                account["statement_balance"] = summary_balance

            # Determine status
            account["status"] = _determine_account_status(account)

        # Calculate totals
        totals = _calculate_totals(bank_accounts)

        return {
            "success": True,
            "data": {
                "accounts": bank_accounts,
                "totals": totals,
                "overall_status": _determine_overall_status(bank_accounts),
            },
            "as_at_date": as_at_date,
            "generated_at": utcnow_iso(),
        }

    except Exception as e:
        logger.exception("Error generating bank recon status: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "as_at_date": as_at_date,
            "generated_at": utcnow_iso(),
        }


def _fetch_bank_accounts(access_token: str, tenant_id: str) -> list[dict]:
    """Fetch bank accounts from Xero."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Accounts",
            headers=headers,
            params={"where": 'Type=="BANK"'},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        accounts = []
        for acc in data.get("Accounts", []):
            accounts.append(
                {
                    "account_id": acc.get("AccountID"),
                    "code": acc.get("Code"),
                    "name": acc.get("Name"),
                    "type": acc.get("Type"),
                    "bank_account_type": acc.get("BankAccountType"),
                    "bank_account_number": acc.get("BankAccountNumber"),
                    "currency": acc.get("CurrencyCode", "AUD"),
                    "statement_balance": None,
                    "book_balance": None,
                    "unreconciled_items": [],
                    "unreconciled_count": 0,
                    "unreconciled_amount": 0,
                    "status": "unknown",
                }
            )

        return accounts

    except requests.RequestException as e:
        logger.warning("Failed to fetch bank accounts: %s", e)
        return []


def _fetch_bank_summary(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> dict[str, float]:
    """Fetch bank summary report for balances."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Reports/BankSummary",
            headers=headers,
            params={"date": as_at_date},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        balances = {}
        reports = data.get("Reports", [])
        if reports:
            for row in reports[0].get("Rows", []):
                if row.get("RowType") == "Section":
                    for section_row in row.get("Rows", []):
                        if section_row.get("RowType") == "Row":
                            cells = section_row.get("Cells", [])
                            if len(cells) >= 2:
                                account_code = cells[0].get("Value", "")
                                # Extract just the account code from "Code - Name" format
                                if " - " in str(account_code):
                                    account_code = account_code.split(" - ")[0].strip()
                                try:
                                    balance = float(
                                        str(cells[1].get("Value", 0))
                                        .replace("$", "")
                                        .replace(",", "")
                                    )
                                    balances[account_code] = balance
                                except (ValueError, TypeError):
                                    pass

        return balances

    except requests.RequestException as e:
        logger.debug("Could not fetch bank summary: %s", e)
        return {}


def _fetch_unreconciled_transactions(
    access_token: str,
    tenant_id: str,
    account_id: str,
) -> list[dict]:
    """Fetch unreconciled bank transactions for an account."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        # Fetch bank transactions that are not reconciled
        resp = requests.get(
            f"{XERO_API_URL}/BankTransactions",
            headers=headers,
            params={
                "where": f'BankAccount.AccountID==guid("{account_id}") AND '
                "IsReconciled==false",
                "order": "Date DESC",
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        transactions = []
        for txn in data.get("BankTransactions", []):
            amount = float(txn.get("Total", 0) or 0)
            if txn.get("Type") == "SPEND":
                amount = -abs(amount)

            transactions.append(
                {
                    "transaction_id": txn.get("BankTransactionID"),
                    "date": txn.get("Date", "")[:10] if txn.get("Date") else "",
                    "type": txn.get("Type"),
                    "contact": txn.get("Contact", {}).get("Name", ""),
                    "reference": txn.get("Reference", ""),
                    "amount": amount,
                    "status": txn.get("Status"),
                }
            )

        return transactions

    except requests.RequestException as e:
        logger.debug("Could not fetch unreconciled transactions: %s", e)
        return []


def _determine_account_status(account: dict) -> str:
    """Determine reconciliation status for an account."""
    unreconciled_count = account.get("unreconciled_count", 0)
    unreconciled_amount = abs(account.get("unreconciled_amount", 0))

    if unreconciled_count == 0:
        return "ok"
    if unreconciled_count <= 5 and unreconciled_amount < 1000:
        return "warning"
    return "error"


def _determine_overall_status(accounts: list[dict]) -> str:
    """Determine overall reconciliation status."""
    if not accounts:
        return "unknown"

    statuses = [acc.get("status") for acc in accounts]

    if all(s == "ok" for s in statuses):
        return "ok"
    if any(s == "error" for s in statuses):
        return "error"
    return "warning"


def _calculate_totals(accounts: list[dict]) -> dict[str, Any]:
    """Calculate totals across all accounts."""
    total_statement = sum(acc.get("statement_balance", 0) or 0 for acc in accounts)
    total_unreconciled_count = sum(acc.get("unreconciled_count", 0) for acc in accounts)
    total_unreconciled_amount = sum(
        acc.get("unreconciled_amount", 0) for acc in accounts
    )

    return {
        "total_bank_balance": round(total_statement, 2),
        "total_unreconciled_count": total_unreconciled_count,
        "total_unreconciled_amount": round(total_unreconciled_amount, 2),
        "account_count": len(accounts),
        "accounts_ok": sum(1 for acc in accounts if acc.get("status") == "ok"),
        "accounts_warning": sum(
            1 for acc in accounts if acc.get("status") == "warning"
        ),
        "accounts_error": sum(1 for acc in accounts if acc.get("status") == "error"),
    }


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export bank recon status to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Recon Status"

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    warning_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )
    error_fill = PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    )

    result = data.get("data", {})

    # Title
    ws["A1"] = "Bank Reconciliation Status"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"As at: {data.get('as_at_date')}"
    ws["A2"].font = Font(italic=True)

    # Summary
    row = 4
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = Font(bold=True)

    totals = result.get("totals", {})
    row += 1
    summary_items = [
        ("Total Bank Balance", totals.get("total_bank_balance", 0)),
        ("Unreconciled Count", totals.get("total_unreconciled_count", 0)),
        ("Unreconciled Amount", totals.get("total_unreconciled_amount", 0)),
        ("Accounts OK", totals.get("accounts_ok", 0)),
        ("Accounts Warning", totals.get("accounts_warning", 0)),
        ("Accounts Error", totals.get("accounts_error", 0)),
    ]

    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if "Amount" in label or "Balance" in label:
            cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Accounts table
    ws[f"A{row}"] = "Bank Accounts"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = [
        "Account",
        "Code",
        "Statement Balance",
        "Unreconciled Items",
        "Unreconciled Amount",
        "Status",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for acc in result.get("accounts", []):
        status = acc.get("status", "unknown")
        fill = (
            ok_fill
            if status == "ok"
            else warning_fill
            if status == "warning"
            else error_fill
        )

        ws.cell(row=row, column=1, value=acc.get("name", ""))
        ws.cell(row=row, column=2, value=acc.get("code", ""))
        ws.cell(
            row=row, column=3, value=acc.get("statement_balance") or 0
        ).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4, value=acc.get("unreconciled_count", 0))
        ws.cell(
            row=row, column=5, value=acc.get("unreconciled_amount", 0)
        ).number_format = '"$"#,##0.00'
        cell = ws.cell(row=row, column=6, value=status.upper())
        cell.fill = fill

        row += 1

    row += 2

    # Unreconciled transactions
    ws[f"A{row}"] = "Unreconciled Transactions"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    txn_headers = ["Account", "Date", "Type", "Contact", "Reference", "Amount"]
    for col, header in enumerate(txn_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for acc in result.get("accounts", []):
        for txn in acc.get("unreconciled_items", []):
            ws.cell(row=row, column=1, value=acc.get("name", ""))
            ws.cell(row=row, column=2, value=txn.get("date", ""))
            ws.cell(row=row, column=3, value=txn.get("type", ""))
            ws.cell(row=row, column=4, value=txn.get("contact", ""))
            ws.cell(row=row, column=5, value=txn.get("reference", ""))
            ws.cell(
                row=row, column=6, value=txn.get("amount", 0)
            ).number_format = '"$"#,##0.00'
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
