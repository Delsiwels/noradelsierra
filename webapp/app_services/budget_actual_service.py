"""
Budget vs Actual Report Service

Compare budgeted vs actual performance with variance analysis.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"


def generate_budget_vs_actual(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
    budget_data: dict[str, float] | None = None,
) -> dict[str, Any]:
    """
    Generate budget vs actual comparison report.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        budget_data: Optional manual budget data

    Returns:
        Dict with budget vs actual comparison
    """
    try:
        # Fetch actual P&L data
        actual_data = _fetch_profit_and_loss(
            access_token, tenant_id, from_date, to_date
        )

        # Try to fetch budget from Xero (may not be available)
        xero_budget = _fetch_xero_budget(access_token, tenant_id, from_date, to_date)

        # Use provided budget data or Xero budget
        budget = budget_data if budget_data else xero_budget

        # Calculate variances
        comparison = _calculate_variances(actual_data, budget)

        # Calculate summary
        summary = _calculate_summary(comparison)

        return {
            "success": True,
            "data": {
                "comparison": comparison,
                "summary": summary,
                "actual": actual_data,
                "budget_source": "manual"
                if budget_data
                else ("xero" if xero_budget else "none"),
            },
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.exception("Error generating budget vs actual: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": datetime.utcnow().isoformat(),
        }


def _fetch_profit_and_loss(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, Any]:
    """Fetch actual P&L data from Xero."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Reports/ProfitAndLoss",
            headers=headers,
            params={"fromDate": from_date, "toDate": to_date},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        return _parse_pnl_report(data)

    except requests.RequestException as e:
        logger.warning("Failed to fetch P&L: %s", e)
        return {"accounts": [], "totals": {}}


def _parse_pnl_report(report_data: dict) -> dict[str, Any]:
    """Parse P&L report into account-level data."""
    reports = report_data.get("Reports", [])
    if not reports:
        return {"accounts": [], "totals": {}}

    accounts = []
    section_totals = {}

    for row in reports[0].get("Rows", []):
        row_type = row.get("RowType")

        if row_type == "Section":
            title = row.get("Title", "")
            section_type = (
                "revenue"
                if "Income" in title or "Revenue" in title
                else "expense"
                if "Expense" in title or "Cost" in title
                else "other"
            )

            for section_row in row.get("Rows", []):
                if section_row.get("RowType") == "Row":
                    cells = section_row.get("Cells", [])
                    if len(cells) >= 2:
                        account_name = cells[0].get("Value", "")
                        try:
                            amount = float(
                                str(cells[1].get("Value", 0))
                                .replace("$", "")
                                .replace(",", "")
                            )
                        except (ValueError, TypeError):
                            amount = 0

                        if account_name and account_name != "Total":
                            accounts.append(
                                {
                                    "account": account_name,
                                    "section": section_type,
                                    "actual": round(amount, 2),
                                }
                            )

                elif section_row.get("RowType") == "SummaryRow":
                    cells = section_row.get("Cells", [])
                    if len(cells) >= 2:
                        try:
                            total = float(
                                str(cells[1].get("Value", 0))
                                .replace("$", "")
                                .replace(",", "")
                            )
                            section_totals[section_type] = round(total, 2)
                        except (ValueError, TypeError):
                            pass

    return {
        "accounts": accounts,
        "totals": section_totals,
    }


def _fetch_xero_budget(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, float] | None:
    """
    Attempt to fetch budget data from Xero.

    Note: Xero Budgets API has limited availability.
    Returns None if not available.
    """
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        # Try Budget Summary report
        resp = requests.get(
            f"{XERO_API_URL}/Reports/BudgetSummary",
            headers=headers,
            params={"date": to_date},
            timeout=30,
        )

        if resp.status_code == 200:
            data = resp.json()
            # Parse budget data if available
            return _parse_budget_report(data)

    except requests.RequestException as e:
        logger.debug("Budget data not available: %s", e)

    return None


def _parse_budget_report(report_data: dict) -> dict[str, float] | None:
    """Parse budget report into account-level data."""
    reports = report_data.get("Reports", [])
    if not reports:
        return None

    budget = {}
    for row in reports[0].get("Rows", []):
        if row.get("RowType") == "Section":
            for section_row in row.get("Rows", []):
                if section_row.get("RowType") == "Row":
                    cells = section_row.get("Cells", [])
                    if len(cells) >= 2:
                        account = cells[0].get("Value", "")
                        try:
                            amount = float(
                                str(cells[1].get("Value", 0))
                                .replace("$", "")
                                .replace(",", "")
                            )
                            if account:
                                budget[account] = amount
                        except (ValueError, TypeError):
                            pass

    return budget if budget else None


def _calculate_variances(
    actual_data: dict[str, Any],
    budget: dict[str, float] | None,
) -> list[dict]:
    """Calculate variance for each account."""
    comparison = []

    for acc in actual_data.get("accounts", []):
        account_name = acc.get("account", "")
        actual = acc.get("actual", 0)
        section = acc.get("section", "other")

        # Get budget amount
        budget_amount = budget.get(account_name, 0) if budget else 0

        # Calculate variance
        variance = actual - budget_amount

        # Variance percentage
        variance_pct = 0
        if budget_amount != 0:
            variance_pct = round((variance / abs(budget_amount)) * 100, 1)

        # Determine status
        # For revenue: positive variance is good
        # For expenses: negative variance is good
        if budget_amount == 0:
            status = "no_budget"
        elif section == "revenue":
            status = "favorable" if variance >= 0 else "unfavorable"
        else:  # expense
            status = "favorable" if variance <= 0 else "unfavorable"

        comparison.append(
            {
                "account": account_name,
                "section": section,
                "budget": round(budget_amount, 2),
                "actual": round(actual, 2),
                "variance": round(variance, 2),
                "variance_pct": variance_pct,
                "status": status,
            }
        )

    return comparison


def _calculate_summary(comparison: list[dict]) -> dict[str, Any]:
    """Calculate summary from comparison data."""
    revenue_items = [c for c in comparison if c["section"] == "revenue"]
    expense_items = [c for c in comparison if c["section"] == "expense"]

    budget_revenue = sum(c["budget"] for c in revenue_items)
    actual_revenue = sum(c["actual"] for c in revenue_items)
    budget_expense = sum(c["budget"] for c in expense_items)
    actual_expense = sum(c["actual"] for c in expense_items)

    budget_profit = budget_revenue - budget_expense
    actual_profit = actual_revenue - actual_expense
    profit_variance = actual_profit - budget_profit

    return {
        "budget_revenue": round(budget_revenue, 2),
        "actual_revenue": round(actual_revenue, 2),
        "revenue_variance": round(actual_revenue - budget_revenue, 2),
        "budget_expense": round(budget_expense, 2),
        "actual_expense": round(actual_expense, 2),
        "expense_variance": round(actual_expense - budget_expense, 2),
        "budget_profit": round(budget_profit, 2),
        "actual_profit": round(actual_profit, 2),
        "profit_variance": round(profit_variance, 2),
        "profit_variance_pct": (
            round((profit_variance / abs(budget_profit)) * 100, 1)
            if budget_profit != 0
            else 0
        ),
    }


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export budget vs actual to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget vs Actual"

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    favorable_fill = PatternFill(
        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
    )
    unfavorable_fill = PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    )

    result = data.get("data", {})
    period = data.get("period", {})

    # Title
    ws["A1"] = "Budget vs Actual Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Period: {period.get('from_date')} to {period.get('to_date')}"
    ws["A2"].font = Font(italic=True)

    ws["A3"] = f"Budget Source: {result.get('budget_source', 'none')}"

    # Summary
    row = 5
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = Font(bold=True)

    summary = result.get("summary", {})
    row += 1
    summary_items = [
        ("Budget Revenue", summary.get("budget_revenue", 0)),
        ("Actual Revenue", summary.get("actual_revenue", 0)),
        ("Revenue Variance", summary.get("revenue_variance", 0)),
        ("Budget Expenses", summary.get("budget_expense", 0)),
        ("Actual Expenses", summary.get("actual_expense", 0)),
        ("Expense Variance", summary.get("expense_variance", 0)),
        ("Budget Profit", summary.get("budget_profit", 0)),
        ("Actual Profit", summary.get("actual_profit", 0)),
        ("Profit Variance", summary.get("profit_variance", 0)),
    ]

    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Detail table
    ws[f"A{row}"] = "Account Detail"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = ["Account", "Section", "Budget", "Actual", "Variance", "Variance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for item in result.get("comparison", []):
        status = item.get("status", "")
        fill = (
            favorable_fill
            if status == "favorable"
            else unfavorable_fill
            if status == "unfavorable"
            else None
        )

        ws.cell(row=row, column=1, value=item.get("account", ""))
        ws.cell(row=row, column=2, value=item.get("section", "").title())
        ws.cell(
            row=row, column=3, value=item.get("budget", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=4, value=item.get("actual", 0)
        ).number_format = '"$"#,##0.00'
        cell = ws.cell(row=row, column=5, value=item.get("variance", 0))
        cell.number_format = '"$"#,##0.00'
        if fill:
            cell.fill = fill
        ws.cell(row=row, column=6, value=f"{item.get('variance_pct', 0)}%")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
