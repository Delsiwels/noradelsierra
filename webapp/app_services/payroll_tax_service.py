"""
Payroll Tax Calculator Service

Calculate state payroll tax liabilities using state-specific
thresholds and rates.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

from webapp.time_utils import utcnow_iso

logger = logging.getLogger(__name__)

XERO_PAYROLL_AU_URL = "https://api.xero.com/payroll.xro/1.0"

# State payroll tax thresholds and rates (2024-25)
PAYROLL_TAX_RATES = {
    "NSW": {
        "annual_threshold": 1200000,
        "rate": 0.0545,  # 5.45%
        "monthly_threshold": 100000,
        "description": "NSW Payroll Tax",
    },
    "VIC": {
        "annual_threshold": 900000,
        "rate": 0.0485,  # 4.85% (tiered, simplified)
        "monthly_threshold": 75000,
        "description": "Victoria Payroll Tax",
    },
    "QLD": {
        "annual_threshold": 1300000,
        "rate": 0.0475,  # 4.75%
        "monthly_threshold": 108333,
        "description": "Queensland Payroll Tax",
    },
    "SA": {
        "annual_threshold": 1500000,
        "rate": 0.0495,  # 4.95%
        "monthly_threshold": 125000,
        "description": "South Australia Payroll Tax",
    },
    "WA": {
        "annual_threshold": 1000000,
        "rate": 0.055,  # 5.5%
        "monthly_threshold": 83333,
        "description": "Western Australia Payroll Tax",
    },
    "TAS": {
        "annual_threshold": 1250000,
        "rate": 0.0485,  # 4.85% (tiered, simplified)
        "monthly_threshold": 104167,
        "description": "Tasmania Payroll Tax",
    },
    "NT": {
        "annual_threshold": 1500000,
        "rate": 0.055,  # 5.5%
        "monthly_threshold": 125000,
        "description": "Northern Territory Payroll Tax",
    },
    "ACT": {
        "annual_threshold": 2000000,
        "rate": 0.0685,  # 6.85%
        "monthly_threshold": 166667,
        "description": "ACT Payroll Tax",
    },
}


def calculate_payroll_tax(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
    state: str,
) -> dict[str, Any]:
    """
    Calculate payroll tax for the period.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        state: Australian state code (NSW, VIC, etc.)

    Returns:
        Dict with payroll tax calculation
    """
    try:
        # Validate state
        state_upper = state.upper()
        if state_upper not in PAYROLL_TAX_RATES:
            return {
                "success": False,
                "error": f"Invalid state: {state}. Must be one of {list(PAYROLL_TAX_RATES.keys())}",
                "data": None,
            }

        # Fetch pay runs for the period
        pay_runs = _fetch_pay_runs(access_token, tenant_id, from_date, to_date)

        # Calculate taxable wages
        taxable_wages = _calculate_taxable_wages(pay_runs)

        # Calculate period months
        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
        to_dt = datetime.strptime(to_date, "%Y-%m-%d")
        period_months = (
            (to_dt.year - from_dt.year) * 12 + (to_dt.month - from_dt.month) + 1
        )

        # Calculate payroll tax
        tax_info = PAYROLL_TAX_RATES[state_upper]
        calculation = _calculate_tax(taxable_wages, tax_info, period_months)

        return {
            "success": True,
            "data": {
                "state": state_upper,
                "state_info": tax_info,
                "wages": taxable_wages,
                "calculation": calculation,
                "pay_runs": pay_runs,
            },
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": utcnow_iso(),
        }

    except Exception as e:
        logger.exception("Error calculating payroll tax: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": utcnow_iso(),
        }


def _fetch_pay_runs(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> list[dict]:
    """Fetch pay runs for the period."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_PAYROLL_AU_URL}/PayRuns",
            headers=headers,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
        to_dt = datetime.strptime(to_date, "%Y-%m-%d")

        pay_runs = []
        for pr in data.get("PayRuns", []):
            payment_date = _parse_xero_date(pr.get("PaymentDate"))
            if not payment_date:
                continue

            payment_dt = datetime.strptime(payment_date, "%Y-%m-%d")
            if from_dt <= payment_dt <= to_dt:
                if pr.get("PayRunStatus") == "POSTED":
                    pay_runs.append(
                        {
                            "pay_run_id": pr.get("PayRunID"),
                            "payment_date": payment_date,
                            "gross_wages": float(pr.get("Wages", 0) or 0),
                            "super": float(pr.get("Super", 0) or 0),
                            "employee_count": len(pr.get("Payslips", [])),
                        }
                    )

        return pay_runs

    except requests.RequestException as e:
        logger.warning("Failed to fetch pay runs: %s", e)
        return []


def _calculate_taxable_wages(pay_runs: list[dict]) -> dict[str, float]:
    """Calculate taxable wages from pay runs."""
    gross_wages = sum(pr.get("gross_wages", 0) for pr in pay_runs)

    # Taxable wages typically includes gross wages minus exempt amounts
    # Exemptions vary by state but generally include some super contributions
    # Simplified: taxable = gross wages (actual calculation is more complex)
    taxable_wages = gross_wages

    return {
        "gross_wages": round(gross_wages, 2),
        "exempt_wages": 0,  # Would need detailed payslip data
        "taxable_wages": round(taxable_wages, 2),
        "pay_run_count": len(pay_runs),
    }


def _calculate_tax(
    wages: dict[str, float],
    tax_info: dict[str, Any],
    period_months: int,
) -> dict[str, Any]:
    """Calculate payroll tax liability."""
    taxable_wages = wages.get("taxable_wages", 0)
    annual_threshold = tax_info.get("annual_threshold", 0)
    rate = tax_info.get("rate", 0)

    # Pro-rata threshold for period
    period_threshold = annual_threshold * (period_months / 12)

    # Calculate tax
    if taxable_wages <= period_threshold:
        tax_payable = 0
        wages_over_threshold = 0
    else:
        wages_over_threshold = taxable_wages - period_threshold
        tax_payable = wages_over_threshold * rate

    # Annualize for comparison
    annualized_wages = (
        taxable_wages * (12 / period_months) if period_months < 12 else taxable_wages
    )

    return {
        "taxable_wages": round(taxable_wages, 2),
        "annualized_wages": round(annualized_wages, 2),
        "annual_threshold": annual_threshold,
        "period_threshold": round(period_threshold, 2),
        "rate_percent": round(rate * 100, 2),
        "wages_over_threshold": round(wages_over_threshold, 2),
        "tax_payable": round(tax_payable, 2),
        "threshold_status": "above" if taxable_wages > period_threshold else "below",
        "period_months": period_months,
    }


def _parse_xero_date(date_value: str | None) -> str | None:
    """Parse Xero date format."""
    if not date_value:
        return None

    if "/Date(" in str(date_value):
        try:
            ts = int(
                str(date_value).split("(")[1].split("+")[0].split("-")[0].split(")")[0]
            )
            dt = datetime.fromtimestamp(ts / 1000)
            return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            return None

    return str(date_value)


def get_all_state_rates() -> dict[str, dict]:
    """Return all state payroll tax rates."""
    return PAYROLL_TAX_RATES


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export payroll tax calculation to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Tax"

    result = data.get("data", {})
    period = data.get("period", {})

    # Title
    ws["A1"] = "Payroll Tax Calculation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    ws["A2"] = f"Period: {period.get('from_date')} to {period.get('to_date')}"
    ws["A2"].font = Font(italic=True)

    ws[
        "A3"
    ] = f"State: {result.get('state', '')} - {result.get('state_info', {}).get('description', '')}"

    # Wages Summary
    row = 5
    ws[f"A{row}"] = "Wages Summary"
    ws[f"A{row}"].font = Font(bold=True)

    wages = result.get("wages", {})
    row += 1
    wages_items = [
        ("Gross Wages", wages.get("gross_wages", 0)),
        ("Exempt Wages", wages.get("exempt_wages", 0)),
        ("Taxable Wages", wages.get("taxable_wages", 0)),
        ("Pay Runs", wages.get("pay_run_count", 0)),
    ]

    for label, value in wages_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Tax Calculation
    ws[f"A{row}"] = "Tax Calculation"
    ws[f"A{row}"].font = Font(bold=True)

    calc = result.get("calculation", {})
    row += 1
    calc_items = [
        ("Annual Threshold", calc.get("annual_threshold", 0)),
        ("Period Threshold", calc.get("period_threshold", 0)),
        ("Rate", f"{calc.get('rate_percent', 0)}%"),
        ("Wages Over Threshold", calc.get("wages_over_threshold", 0)),
        ("Tax Payable", calc.get("tax_payable", 0)),
    ]

    for label, value in calc_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, int | float):
            cell.number_format = '"$"#,##0.00'
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
