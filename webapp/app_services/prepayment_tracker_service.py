"""
Prepayment Tracker Service

Track prepaid expenses and amortization schedules.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

from webapp.time_utils import utcnow_iso

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"


def generate_prepayment_schedule(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> dict[str, Any]:
    """
    Generate prepayment tracking schedule.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        as_at_date: Date for analysis (YYYY-MM-DD)

    Returns:
        Dict with prepayment schedule and amortization status
    """
    try:
        # Fetch prepayment accounts
        prepayment_accounts = _fetch_prepayment_accounts(access_token, tenant_id)

        # Fetch journals for movements
        journals = _fetch_prepayment_journals(access_token, tenant_id, as_at_date)

        # Build schedule with movements
        schedule = _build_prepayment_schedule(prepayment_accounts, journals, as_at_date)

        # Calculate totals
        totals = _calculate_totals(schedule)

        return {
            "success": True,
            "data": {
                "schedule": schedule,
                "totals": totals,
            },
            "as_at_date": as_at_date,
            "generated_at": utcnow_iso(),
        }

    except Exception as e:
        logger.exception("Error generating prepayment schedule: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "as_at_date": as_at_date,
            "generated_at": utcnow_iso(),
        }


def _fetch_prepayment_accounts(
    access_token: str,
    tenant_id: str,
) -> list[dict]:
    """Fetch prepayment and prepaid expense accounts."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Accounts",
            headers=headers,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        accounts = []
        for acc in data.get("Accounts", []):
            acc_type = acc.get("Type", "")
            acc_name = acc.get("Name", "").lower()

            # Look for prepayment accounts
            is_prepayment = (
                acc_type == "PREPAYMENT"
                or "prepaid" in acc_name
                or "prepayment" in acc_name
            )

            if is_prepayment:
                accounts.append(
                    {
                        "account_id": acc.get("AccountID"),
                        "code": acc.get("Code"),
                        "name": acc.get("Name"),
                        "type": acc_type,
                    }
                )

        return accounts

    except requests.RequestException as e:
        logger.warning("Failed to fetch prepayment accounts: %s", e)
        return []


def _fetch_prepayment_journals(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> list[dict]:
    """Fetch journal entries for prepayment accounts."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Journals",
            headers=headers,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        journals = []
        for journal in data.get("Journals", []):
            journal_date = journal.get("JournalDate", "")[:10]

            for line in journal.get("JournalLines", []):
                journals.append(
                    {
                        "journal_id": journal.get("JournalID"),
                        "date": journal_date,
                        "account_code": line.get("AccountCode"),
                        "account_name": line.get("AccountName"),
                        "debit": (
                            float(line.get("GrossAmount", 0) or 0)
                            if float(line.get("GrossAmount", 0) or 0) > 0
                            else 0
                        ),
                        "credit": (
                            abs(float(line.get("GrossAmount", 0) or 0))
                            if float(line.get("GrossAmount", 0) or 0) < 0
                            else 0
                        ),
                        "description": line.get("Description", ""),
                    }
                )

        return journals

    except requests.RequestException as e:
        logger.debug("Could not fetch journals: %s", e)
        return []


def _build_prepayment_schedule(
    accounts: list[dict],
    journals: list[dict],
    as_at_date: str,
) -> list[dict]:
    """Build prepayment schedule with movements."""
    schedule = []

    as_at_dt = datetime.strptime(as_at_date, "%Y-%m-%d")

    # Determine FY start for opening balance
    if as_at_dt.month >= 7:
        fy_start_dt = datetime(as_at_dt.year, 7, 1)
    else:
        fy_start_dt = datetime(as_at_dt.year - 1, 7, 1)

    for account in accounts:
        account_code = account.get("code", "")

        # Filter journals for this account
        account_journals = [
            j for j in journals if j.get("account_code") == account_code
        ]

        # Calculate opening (before FY start)
        opening = 0
        additions = 0
        amortization = 0

        for j in account_journals:
            try:
                journal_dt = datetime.strptime(j.get("date", ""), "%Y-%m-%d")
            except ValueError:
                continue

            if journal_dt < fy_start_dt:
                opening += j.get("debit", 0) - j.get("credit", 0)
            elif journal_dt <= as_at_dt:
                debit = j.get("debit", 0)
                credit = j.get("credit", 0)
                if debit > 0:
                    additions += debit
                if credit > 0:
                    amortization += credit

        closing = opening + additions - amortization

        # Estimate monthly amortization (if balance exists)
        monthly_amortization: float = 0.0
        if closing > 0 and amortization > 0:
            months_elapsed = (
                (as_at_dt.year - fy_start_dt.year) * 12
                + (as_at_dt.month - fy_start_dt.month)
                + 1
            )
            if months_elapsed > 0:
                monthly_amortization = amortization / months_elapsed

        # Determine status
        status = "ok"
        if closing > 0 and amortization == 0:
            status = "warning"  # Prepayment exists but no amortization recorded
        elif closing < 0:
            status = "error"  # Negative balance

        schedule.append(
            {
                "account_code": account_code,
                "account_name": account.get("name", ""),
                "opening": round(opening, 2),
                "additions": round(additions, 2),
                "amortization": round(amortization, 2),
                "closing": round(closing, 2),
                "monthly_amortization": round(monthly_amortization, 2),
                "status": status,
            }
        )

    return schedule


def _calculate_totals(schedule: list[dict]) -> dict[str, Any]:
    """Calculate totals from schedule."""
    return {
        "total_opening": round(sum(s.get("opening", 0) for s in schedule), 2),
        "total_additions": round(sum(s.get("additions", 0) for s in schedule), 2),
        "total_amortization": round(sum(s.get("amortization", 0) for s in schedule), 2),
        "total_closing": round(sum(s.get("closing", 0) for s in schedule), 2),
        "total_monthly": round(
            sum(s.get("monthly_amortization", 0) for s in schedule), 2
        ),
        "account_count": len(schedule),
        "accounts_ok": sum(1 for s in schedule if s.get("status") == "ok"),
        "accounts_warning": sum(1 for s in schedule if s.get("status") == "warning"),
    }


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export prepayment schedule to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prepayment Tracker"

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    warning_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )

    result = data.get("data", {})

    # Title
    ws["A1"] = "Prepayment Tracker"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    ws["A2"] = f"As at: {data.get('as_at_date')}"
    ws["A2"].font = Font(italic=True)

    # Summary
    row = 4
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = Font(bold=True)

    totals = result.get("totals", {})
    row += 1
    summary_items = [
        ("Total Opening", totals.get("total_opening", 0)),
        ("Total Additions", totals.get("total_additions", 0)),
        ("Total Amortization", totals.get("total_amortization", 0)),
        ("Total Closing", totals.get("total_closing", 0)),
        ("Est. Monthly Amortization", totals.get("total_monthly", 0)),
    ]

    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Schedule table
    ws[f"A{row}"] = "Prepayment Schedule"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = [
        "Account",
        "Opening",
        "Additions",
        "Amortization",
        "Closing",
        "Monthly Est.",
        "Status",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for item in result.get("schedule", []):
        status = item.get("status", "")

        ws.cell(row=row, column=1, value=item.get("account_name", ""))
        ws.cell(
            row=row, column=2, value=item.get("opening", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=3, value=item.get("additions", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=4, value=item.get("amortization", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=item.get("closing", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=6, value=item.get("monthly_amortization", 0)
        ).number_format = '"$"#,##0.00'
        cell = ws.cell(row=row, column=7, value=status.upper())
        if status == "warning":
            cell.fill = warning_fill
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
