"""
AR/AP Aging Dashboard Service

Interactive aging analysis with overdue alerts for accounts receivable
and accounts payable.
"""

import logging
from io import BytesIO
from typing import Any

import requests

from webapp.time_utils import utcnow_iso

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"


def generate_aging_dashboard(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> dict[str, Any]:
    """
    Generate AR/AP aging dashboard data.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        as_at_date: Date for aging calculation (YYYY-MM-DD)

    Returns:
        Dict with receivables, payables, summaries, and alerts
    """
    try:
        # Fetch aged receivables
        receivables = _fetch_aged_receivables(access_token, tenant_id, as_at_date)

        # Fetch aged payables
        payables = _fetch_aged_payables(access_token, tenant_id, as_at_date)

        # Calculate summaries
        ar_summary = _calculate_summary(receivables)
        ap_summary = _calculate_summary(payables)

        # Generate overdue alerts
        ar_alerts = _generate_alerts(receivables, "receivable")
        ap_alerts = _generate_alerts(payables, "payable")

        return {
            "success": True,
            "data": {
                "receivables": receivables,
                "payables": payables,
                "ar_summary": ar_summary,
                "ap_summary": ap_summary,
                "ar_alerts": ar_alerts,
                "ap_alerts": ap_alerts,
            },
            "as_at_date": as_at_date,
            "generated_at": utcnow_iso(),
        }

    except Exception as e:
        logger.exception("Error generating aging dashboard: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "as_at_date": as_at_date,
            "generated_at": utcnow_iso(),
        }


def _fetch_aged_receivables(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> list[dict]:
    """Fetch aged receivables by contact."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Reports/AgedReceivablesByContact",
            headers=headers,
            params={"date": as_at_date},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        return _parse_aged_report(data)

    except requests.RequestException as e:
        logger.warning("Failed to fetch aged receivables: %s", e)
        return []


def _fetch_aged_payables(
    access_token: str,
    tenant_id: str,
    as_at_date: str,
) -> list[dict]:
    """Fetch aged payables by contact."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Reports/AgedPayablesByContact",
            headers=headers,
            params={"date": as_at_date},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        return _parse_aged_report(data)

    except requests.RequestException as e:
        logger.warning("Failed to fetch aged payables: %s", e)
        return []


def _parse_aged_report(report_data: dict) -> list[dict]:
    """Parse Xero aged report format into structured data."""
    reports = report_data.get("Reports", [])
    if not reports:
        return []

    report = reports[0]
    rows = report.get("Rows", [])

    contacts = []
    header_found = False

    for row in rows:
        row_type = row.get("RowType")

        if row_type == "Header":
            header_found = True
            continue

        if row_type == "Section":
            section_rows = row.get("Rows", [])
            for section_row in section_rows:
                if section_row.get("RowType") == "Row":
                    contact_data = _parse_contact_row(section_row)
                    if contact_data:
                        contacts.append(contact_data)

        elif row_type == "Row" and header_found:
            contact_data = _parse_contact_row(row)
            if contact_data:
                contacts.append(contact_data)

    return contacts


def _parse_contact_row(row: dict) -> dict | None:
    """Parse a single contact row from aged report."""
    cells = row.get("Cells", [])
    if len(cells) < 6:
        return None

    try:
        contact_name = cells[0].get("Value", "")
        if not contact_name or contact_name == "Total":
            return None

        # Extract amounts from cells
        # Typical structure: Contact, Current, 30 days, 60 days, 90+ days, Total
        current = _parse_amount(cells[1].get("Value"))
        days_30 = _parse_amount(cells[2].get("Value"))
        days_60 = _parse_amount(cells[3].get("Value"))
        days_90_plus = _parse_amount(cells[4].get("Value"))
        total = _parse_amount(cells[5].get("Value"))

        # If total is 0 or None, skip
        if not total:
            return None

        return {
            "contact_name": contact_name,
            "contact_id": (
                cells[0].get("Attributes", [{}])[0].get("Value")
                if cells[0].get("Attributes")
                else None
            ),
            "current": current,
            "days_30": days_30,
            "days_60": days_60,
            "days_90_plus": days_90_plus,
            "total": total,
            "overdue_60_plus": (days_60 or 0) + (days_90_plus or 0),
        }

    except (IndexError, KeyError, TypeError) as e:
        logger.debug("Error parsing contact row: %s", e)
        return None


def _parse_amount(value: Any) -> float:
    """Parse amount value, handling various formats."""
    if value is None or value == "":
        return 0.0
    try:
        # Remove currency symbols and commas
        cleaned = str(value).replace("$", "").replace(",", "").strip()
        return float(cleaned)
    except ValueError:
        return 0.0


def _calculate_summary(contacts: list[dict]) -> dict[str, Any]:
    """Calculate summary totals from contact list."""
    total = sum(c.get("total", 0) for c in contacts)
    current = sum(c.get("current", 0) for c in contacts)
    days_30 = sum(c.get("days_30", 0) for c in contacts)
    days_60 = sum(c.get("days_60", 0) for c in contacts)
    days_90_plus = sum(c.get("days_90_plus", 0) for c in contacts)
    overdue_60_plus = sum(c.get("overdue_60_plus", 0) for c in contacts)

    return {
        "total": round(total, 2),
        "current": round(current, 2),
        "days_30": round(days_30, 2),
        "days_60": round(days_60, 2),
        "days_90_plus": round(days_90_plus, 2),
        "overdue_60_plus": round(overdue_60_plus, 2),
        "contact_count": len(contacts),
        "overdue_contact_count": sum(
            1 for c in contacts if c.get("overdue_60_plus", 0) > 0
        ),
        "current_pct": round(current / total * 100, 1) if total else 0,
        "days_30_pct": round(days_30 / total * 100, 1) if total else 0,
        "days_60_pct": round(days_60 / total * 100, 1) if total else 0,
        "days_90_plus_pct": round(days_90_plus / total * 100, 1) if total else 0,
    }


def _generate_alerts(
    contacts: list[dict],
    alert_type: str,
    threshold: float = 500.0,
) -> list[dict]:
    """Generate alerts for overdue contacts."""
    alerts = []

    for contact in contacts:
        overdue_60_plus = contact.get("overdue_60_plus", 0)
        days_90_plus = contact.get("days_90_plus", 0)

        if overdue_60_plus >= threshold:
            severity = "high" if days_90_plus >= threshold else "medium"

            alerts.append(
                {
                    "contact_name": contact.get("contact_name"),
                    "contact_id": contact.get("contact_id"),
                    "type": alert_type,
                    "severity": severity,
                    "amount": overdue_60_plus,
                    "days_90_plus": days_90_plus,
                    "message": (
                        f"${overdue_60_plus:,.2f} overdue 60+ days"
                        if severity == "medium"
                        else f"${days_90_plus:,.2f} overdue 90+ days"
                    ),
                }
            )

    # Sort by amount descending
    alerts.sort(key=lambda x: x.get("amount", 0), reverse=True)
    return alerts


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export aging dashboard to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter  # noqa: F401
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    alert_fill = PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    )
    warning_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )

    result = data.get("data", {})
    as_at_date = str(data.get("as_at_date", ""))

    # Receivables sheet
    ws_ar = wb.active
    ws_ar.title = "Receivables"
    _write_aging_sheet(
        ws_ar,
        result.get("receivables", []),
        result.get("ar_summary", {}),
        "Aged Receivables",
        as_at_date,
        header_fill,
        header_font,
        alert_fill,
        warning_fill,
    )

    # Payables sheet
    ws_ap = wb.create_sheet("Payables")
    _write_aging_sheet(
        ws_ap,
        result.get("payables", []),
        result.get("ap_summary", {}),
        "Aged Payables",
        as_at_date,
        header_fill,
        header_font,
        alert_fill,
        warning_fill,
    )

    # Alerts sheet
    ws_alerts = wb.create_sheet("Alerts")
    _write_alerts_sheet(
        ws_alerts,
        result.get("ar_alerts", []) + result.get("ap_alerts", []),
        header_fill,
        header_font,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_aging_sheet(
    ws,
    contacts: list[dict],
    summary: dict,
    title: str,
    as_at_date: str,
    header_fill,
    header_font,
    alert_fill,
    warning_fill,
):
    """Write aging data to a worksheet."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    # Title
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"As at: {as_at_date}"
    ws["A2"].font = Font(italic=True)

    # Summary
    row = 4
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    summary_items = [
        ("Total Outstanding", summary.get("total", 0)),
        ("Current", summary.get("current", 0)),
        ("30 Days", summary.get("days_30", 0)),
        ("60 Days", summary.get("days_60", 0)),
        ("90+ Days", summary.get("days_90_plus", 0)),
        ("Overdue (60+)", summary.get("overdue_60_plus", 0)),
    ]

    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Detail table
    ws[f"A{row}"] = "By Contact"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = ["Contact", "Current", "30 Days", "60 Days", "90+ Days", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for contact in contacts:
        ws.cell(row=row, column=1, value=contact.get("contact_name", ""))
        ws.cell(
            row=row, column=2, value=contact.get("current", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=3, value=contact.get("days_30", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=4, value=contact.get("days_60", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=contact.get("days_90_plus", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=6, value=contact.get("total", 0)
        ).number_format = '"$"#,##0.00'

        # Highlight overdue rows
        if contact.get("days_90_plus", 0) > 500:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = alert_fill
        elif contact.get("days_60", 0) > 500:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = warning_fill

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _write_alerts_sheet(ws, alerts: list[dict], header_fill, header_font):
    """Write alerts to a worksheet."""
    from openpyxl.styles import Font

    ws["A1"] = "Overdue Alerts"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    headers = ["Contact", "Type", "Severity", "Amount", "90+ Days", "Message"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for alert in alerts:
        ws.cell(row=row, column=1, value=alert.get("contact_name", ""))
        ws.cell(row=row, column=2, value=alert.get("type", "").title())
        ws.cell(row=row, column=3, value=alert.get("severity", "").upper())
        ws.cell(
            row=row, column=4, value=alert.get("amount", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=alert.get("days_90_plus", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=alert.get("message", ""))
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 35
