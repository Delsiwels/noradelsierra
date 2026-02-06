"""
STP Submission Tracker Service

Track STP lodgement status and payroll totals by quarter.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

logger = logging.getLogger(__name__)

XERO_PAYROLL_AU_URL = "https://api.xero.com/payroll.xro/1.0"


def generate_stp_summary(
    access_token: str,
    tenant_id: str,
    financial_year: int,
) -> dict[str, Any]:
    """
    Generate STP summary for a financial year.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        financial_year: Financial year (e.g., 2025 for FY2024-25)

    Returns:
        Dict with STP summary by quarter
    """
    try:
        # Calculate FY date range
        fy_start = f"{financial_year - 1}-07-01"
        fy_end = f"{financial_year}-06-30"

        # Fetch all pay runs for the FY
        pay_runs = _fetch_pay_runs_for_fy(access_token, tenant_id, fy_start, fy_end)

        # Group by quarter
        quarters = _group_by_quarter(pay_runs, financial_year)

        # Calculate YTD totals
        ytd_totals = _calculate_ytd_totals(quarters)

        return {
            "success": True,
            "data": {
                "quarters": quarters,
                "ytd_totals": ytd_totals,
                "pay_runs": pay_runs,
            },
            "financial_year": f"FY{financial_year - 1}-{str(financial_year)[-2:]}",
            "generated_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.exception("Error generating STP summary: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "financial_year": f"FY{financial_year - 1}-{str(financial_year)[-2:]}",
            "generated_at": datetime.utcnow().isoformat(),
        }


def _fetch_pay_runs_for_fy(
    access_token: str,
    tenant_id: str,
    fy_start: str,
    fy_end: str,
) -> list[dict]:
    """Fetch all pay runs for the financial year."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_PAYROLL_AU_URL}/PayRuns",
            headers=headers,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        fy_start_dt = datetime.strptime(fy_start, "%Y-%m-%d")
        fy_end_dt = datetime.strptime(fy_end, "%Y-%m-%d")

        pay_runs = []
        for pr in data.get("PayRuns", []):
            payment_date = _parse_xero_date(pr.get("PaymentDate"))
            if not payment_date:
                continue

            payment_dt = datetime.strptime(payment_date, "%Y-%m-%d")
            if fy_start_dt <= payment_dt <= fy_end_dt:
                # Only include POSTED pay runs for STP tracking
                if pr.get("PayRunStatus") == "POSTED":
                    pay_runs.append(
                        {
                            "pay_run_id": pr.get("PayRunID"),
                            "payment_date": payment_date,
                            "period_start": _parse_xero_date(
                                pr.get("PayRunPeriodStartDate")
                            ),
                            "period_end": _parse_xero_date(
                                pr.get("PayRunPeriodEndDate")
                            ),
                            "status": pr.get("PayRunStatus"),
                            "gross_wages": float(pr.get("Wages", 0) or 0),
                            "payg_withheld": float(pr.get("Tax", 0) or 0),
                            "super": float(pr.get("Super", 0) or 0),
                            "net_pay": float(pr.get("NetPay", 0) or 0),
                            "employee_count": len(pr.get("Payslips", [])),
                        }
                    )

        # Sort by payment date
        pay_runs.sort(key=lambda x: x.get("payment_date", ""))
        return pay_runs

    except requests.RequestException as e:
        logger.warning("Failed to fetch pay runs: %s", e)
        return []


def _group_by_quarter(pay_runs: list[dict], financial_year: int) -> list[dict]:
    """Group pay runs by quarter."""
    quarters = [
        {
            "quarter": "Q1",
            "period": f"Jul-Sep {financial_year - 1}",
            "start_date": f"{financial_year - 1}-07-01",
            "end_date": f"{financial_year - 1}-09-30",
            "gross_wages": 0,
            "payg_withheld": 0,
            "super": 0,
            "pay_run_count": 0,
            "employee_count": 0,
        },
        {
            "quarter": "Q2",
            "period": f"Oct-Dec {financial_year - 1}",
            "start_date": f"{financial_year - 1}-10-01",
            "end_date": f"{financial_year - 1}-12-31",
            "gross_wages": 0,
            "payg_withheld": 0,
            "super": 0,
            "pay_run_count": 0,
            "employee_count": 0,
        },
        {
            "quarter": "Q3",
            "period": f"Jan-Mar {financial_year}",
            "start_date": f"{financial_year}-01-01",
            "end_date": f"{financial_year}-03-31",
            "gross_wages": 0,
            "payg_withheld": 0,
            "super": 0,
            "pay_run_count": 0,
            "employee_count": 0,
        },
        {
            "quarter": "Q4",
            "period": f"Apr-Jun {financial_year}",
            "start_date": f"{financial_year}-04-01",
            "end_date": f"{financial_year}-06-30",
            "gross_wages": 0,
            "payg_withheld": 0,
            "super": 0,
            "pay_run_count": 0,
            "employee_count": 0,
        },
    ]

    for pr in pay_runs:
        payment_date = pr.get("payment_date", "")
        if not payment_date:
            continue

        payment_dt = datetime.strptime(payment_date, "%Y-%m-%d")

        for q in quarters:
            q_start = datetime.strptime(str(q["start_date"]), "%Y-%m-%d")
            q_end = datetime.strptime(str(q["end_date"]), "%Y-%m-%d")

            if q_start <= payment_dt <= q_end:
                gross: float = float(q.get("gross_wages") or 0)  # type: ignore[arg-type]
                payg: float = float(q.get("payg_withheld") or 0)  # type: ignore[arg-type]
                super_val: float = float(q.get("super") or 0)  # type: ignore[arg-type]
                q["gross_wages"] = gross + float(pr.get("gross_wages") or 0)
                q["payg_withheld"] = payg + float(pr.get("payg_withheld") or 0)
                q["super"] = super_val + float(pr.get("super") or 0)
                cnt: int = int(q.get("pay_run_count") or 0)  # type: ignore[call-overload]
                q["pay_run_count"] = cnt + 1
                emp_q: int = int(q.get("employee_count") or 0)  # type: ignore[call-overload]
                emp_pr: int = int(pr.get("employee_count") or 0)
                q["employee_count"] = max(emp_q, emp_pr)
                break

    # Round values
    for q in quarters:
        g: float = float(q.get("gross_wages") or 0)  # type: ignore[arg-type]
        p: float = float(q.get("payg_withheld") or 0)  # type: ignore[arg-type]
        s: float = float(q.get("super") or 0)  # type: ignore[arg-type]
        q["gross_wages"] = round(g, 2)
        q["payg_withheld"] = round(p, 2)
        q["super"] = round(s, 2)

    return quarters


def _calculate_ytd_totals(quarters: list[dict]) -> dict[str, Any]:
    """Calculate YTD totals from quarters."""
    return {
        "gross_wages": round(sum(q["gross_wages"] for q in quarters), 2),
        "payg_withheld": round(sum(q["payg_withheld"] for q in quarters), 2),
        "super": round(sum(q["super"] for q in quarters), 2),
        "pay_run_count": sum(q["pay_run_count"] for q in quarters),
        "max_employees": max((q["employee_count"] for q in quarters), default=0),
    }


def _parse_xero_date(date_value: str | None) -> str | None:
    """Parse Xero date format /Date(timestamp)/ to YYYY-MM-DD."""
    if not date_value:
        return None

    if "/Date(" in str(date_value):
        try:
            ts = int(
                str(date_value).split("(")[1].split("+")[0].split("-")[0].split(")")[0]
            )
            dt = datetime.fromtimestamp(ts / 1000)
            return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            return None

    return str(date_value)


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export STP summary to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "STP Summary"

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    result = data.get("data", {})

    # Title
    ws["A1"] = "STP Submission Tracker"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Financial Year: {data.get('financial_year')}"
    ws["A2"].font = Font(italic=True)

    # YTD Summary
    row = 4
    ws[f"A{row}"] = "YTD Summary"
    ws[f"A{row}"].font = Font(bold=True)

    ytd = result.get("ytd_totals", {})
    row += 1
    ytd_items = [
        ("Gross Wages", ytd.get("gross_wages", 0)),
        ("PAYG Withheld", ytd.get("payg_withheld", 0)),
        ("Super Guarantee", ytd.get("super", 0)),
        ("Pay Runs Processed", ytd.get("pay_run_count", 0)),
        ("Max Employees", ytd.get("max_employees", 0)),
    ]

    for label, value in ytd_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Quarterly breakdown
    ws[f"A{row}"] = "Quarterly Breakdown"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = ["Quarter", "Period", "Gross Wages", "PAYG Withheld", "Super", "Pay Runs"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for q in result.get("quarters", []):
        ws.cell(row=row, column=1, value=q.get("quarter", ""))
        ws.cell(row=row, column=2, value=q.get("period", ""))
        ws.cell(
            row=row, column=3, value=q.get("gross_wages", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=4, value=q.get("payg_withheld", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=q.get("super", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=q.get("pay_run_count", 0))
        row += 1

    row += 2

    # Note about STP submission status
    ws.cell(
        row=row,
        column=1,
        value="Note: STP submission status is not available via Xero API. "
        "Please check Xero Payroll directly for lodgement confirmation.",
    )
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].font = Font(italic=True, color="666666")

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
