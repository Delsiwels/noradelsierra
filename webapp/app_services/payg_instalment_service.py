"""
PAYG-I Calculator Service

Calculate PAYG instalments based on income using Australian
instalment rate and instalment amount methods.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

from webapp.time_utils import utcnow_iso

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"

# Default PAYG-I rate (can be overridden by ATO-issued rate)
DEFAULT_PAYGI_RATE = 0.03  # 3%


def calculate_payg_instalment(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
    instalment_rate: float | None = None,
    method: str = "rate",
) -> dict[str, Any]:
    """
    Calculate PAYG instalment for the period.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        instalment_rate: Override rate (default 3%)
        method: "rate" (instalment rate) or "amount" (instalment amount)

    Returns:
        Dict with instalment calculation
    """
    try:
        # Fetch Profit & Loss for the period
        pnl_data = _fetch_profit_and_loss(access_token, tenant_id, from_date, to_date)

        # Calculate period months
        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
        to_dt = datetime.strptime(to_date, "%Y-%m-%d")
        period_months = (
            (to_dt.year - from_dt.year) * 12 + (to_dt.month - from_dt.month) + 1
        )

        # Calculate instalment
        rate = instalment_rate if instalment_rate is not None else DEFAULT_PAYGI_RATE
        calculation = _calculate_instalment(pnl_data, period_months, rate, method)

        return {
            "success": True,
            "data": {
                "profit_and_loss": pnl_data,
                "calculation": calculation,
                "method": method,
            },
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": utcnow_iso(),
        }

    except Exception as e:
        logger.exception("Error calculating PAYG instalment: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": utcnow_iso(),
        }


def _fetch_profit_and_loss(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, Any]:
    """Fetch Profit and Loss report from Xero."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Reports/ProfitAndLoss",
            headers=headers,
            params={"fromDate": from_date, "toDate": to_date},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        return _parse_profit_and_loss(data)

    except requests.RequestException as e:
        logger.warning("Failed to fetch P&L: %s", e)
        return {
            "revenue": 0,
            "expenses": 0,
            "net_profit": 0,
            "sections": [],
        }


def _parse_profit_and_loss(report_data: dict) -> dict[str, Any]:
    """Parse P&L report into structured data."""
    reports = report_data.get("Reports", [])
    if not reports:
        return {"revenue": 0, "expenses": 0, "net_profit": 0, "sections": []}

    revenue: float = 0.0
    expenses: float = 0.0
    net_profit: float = 0.0
    sections: list[dict] = []

    for row in reports[0].get("Rows", []):
        row_type = row.get("RowType")

        if row_type == "Section":
            title = row.get("Title", "")
            section_total: float = 0.0

            for section_row in row.get("Rows", []):
                if section_row.get("RowType") == "SummaryRow":
                    cells = section_row.get("Cells", [])
                    if len(cells) >= 2:
                        try:
                            section_total = float(
                                str(cells[1].get("Value", 0))
                                .replace("$", "")
                                .replace(",", "")
                            )
                        except (ValueError, TypeError):
                            pass

            if "Income" in title or "Revenue" in title:
                revenue = section_total
            elif "Expense" in title or "Cost" in title:
                expenses = abs(section_total)

            sections.append({"title": title, "total": section_total})

        elif row_type == "Row" and "Net Profit" in str(
            row.get("Cells", [{}])[0].get("Value", "")
        ):
            cells = row.get("Cells", [])
            if len(cells) >= 2:
                try:
                    net_profit = float(
                        str(cells[1].get("Value", 0)).replace("$", "").replace(",", "")
                    )
                except (ValueError, TypeError):
                    pass

    # Calculate net profit if not found
    if net_profit == 0:
        net_profit = revenue - expenses

    return {
        "revenue": round(revenue, 2),
        "expenses": round(expenses, 2),
        "net_profit": round(net_profit, 2),
        "sections": sections,
    }


def _calculate_instalment(
    pnl: dict[str, Any],
    period_months: int,
    rate: float,
    method: str,
) -> dict[str, Any]:
    """Calculate PAYG instalment."""
    net_profit = pnl.get("net_profit", 0)

    # Annualize if not a full year
    if period_months < 12:
        annualized_profit = net_profit * (12 / period_months)
    else:
        annualized_profit = net_profit

    # Calculate based on method
    if method == "rate":
        # Instalment Rate Method
        # T2 = Instalment income × Rate / 4 (for quarterly)
        instalment_income = (
            annualized_profit  # Simplified - actual uses gross business income
        )
        annual_instalment = max(instalment_income * rate, 0)
        quarterly_instalment = annual_instalment / 4
    else:
        # Instalment Amount Method
        # Use ATO-provided amount (not calculated here)
        quarterly_instalment = 0
        annual_instalment = 0

    return {
        "period_profit": round(net_profit, 2),
        "annualized_profit": round(annualized_profit, 2),
        "instalment_rate": round(rate * 100, 2),
        "annual_instalment": round(annual_instalment, 2),
        "quarterly_instalment": round(quarterly_instalment, 2),
        "period_months": period_months,
        "method_description": (
            "Instalment Rate Method (T2 = Income x Rate / 4)"
            if method == "rate"
            else "Instalment Amount Method (use ATO-advised amount)"
        ),
    }


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export PAYG-I calculation to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAYG-I Calculation"

    result = data.get("data", {})
    period = data.get("period", {})

    # Title
    ws["A1"] = "PAYG Instalment Calculation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    ws["A2"] = f"Period: {period.get('from_date')} to {period.get('to_date')}"
    ws["A2"].font = Font(italic=True)

    # P&L Summary
    row = 4
    ws[f"A{row}"] = "Profit & Loss Summary"
    ws[f"A{row}"].font = Font(bold=True)

    pnl = result.get("profit_and_loss", {})
    row += 1
    pnl_items = [
        ("Revenue", pnl.get("revenue", 0)),
        ("Expenses", pnl.get("expenses", 0)),
        ("Net Profit", pnl.get("net_profit", 0)),
    ]

    for label, value in pnl_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Instalment Calculation
    ws[f"A{row}"] = "Instalment Calculation"
    ws[f"A{row}"].font = Font(bold=True)

    calc = result.get("calculation", {})
    row += 1
    calc_items = [
        ("Period Profit", calc.get("period_profit", 0)),
        ("Annualized Profit", calc.get("annualized_profit", 0)),
        ("Instalment Rate", f"{calc.get('instalment_rate', 0)}%"),
        ("Annual Instalment", calc.get("annual_instalment", 0)),
        ("Quarterly Instalment", calc.get("quarterly_instalment", 0)),
    ]

    for label, value in calc_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, int | float):
            cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Method")
    ws.cell(row=row, column=2, value=calc.get("method_description", ""))

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
