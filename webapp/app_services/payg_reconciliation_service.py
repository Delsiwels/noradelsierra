"""
PAYG-W Reconciliation Service

Reconciles PAYG withholding from payroll against W2 label on BAS.
Compares payroll totals with BAS totals to identify discrepancies.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

logger = logging.getLogger(__name__)

# Xero API base URLs
XERO_PAYROLL_AU_URL = "https://api.xero.com/payroll.xro/1.0"
XERO_API_URL = "https://api.xero.com/api.xro/2.0"


def generate_payg_reconciliation(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, Any]:
    """
    Generate PAYG-W reconciliation report.

    Compares:
    - Payroll PAYG withheld (sum from pay runs)
    - BAS W2 label (PAYG withheld reported on BAS)

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)

    Returns:
        Dict with reconciliation data, warnings, and status
    """
    try:
        # Fetch pay runs for the period
        pay_runs = _fetch_pay_runs(access_token, tenant_id, from_date, to_date)

        # Calculate payroll totals
        payroll_totals = _calculate_payroll_totals(pay_runs)

        # Fetch BAS data if available (from reports or manual entry)
        bas_data = _fetch_bas_data(access_token, tenant_id, from_date, to_date)

        # Calculate variance
        variance = _calculate_variance(payroll_totals, bas_data)

        # Determine status
        status = _determine_status(variance)

        return {
            "success": True,
            "data": {
                "payroll": payroll_totals,
                "bas": bas_data,
                "variance": variance,
                "pay_runs": pay_runs,
                "status": status,
            },
            "warnings": _generate_warnings(variance, pay_runs),
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.exception("Error generating PAYG reconciliation: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "warnings": [],
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": datetime.utcnow().isoformat(),
        }


def _fetch_pay_runs(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> list[dict]:
    """Fetch pay runs within the date range."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_PAYROLL_AU_URL}/PayRuns",
            headers=headers,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        pay_runs = []
        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
        to_dt = datetime.strptime(to_date, "%Y-%m-%d")

        for pr in data.get("PayRuns", []):
            payment_date = _parse_xero_date(pr.get("PaymentDate"))
            if not payment_date:
                continue

            payment_dt = datetime.strptime(payment_date, "%Y-%m-%d")
            if from_dt <= payment_dt <= to_dt:
                pay_runs.append(
                    {
                        "pay_run_id": pr.get("PayRunID"),
                        "payment_date": payment_date,
                        "period_start": _parse_xero_date(
                            pr.get("PayRunPeriodStartDate")
                        ),
                        "period_end": _parse_xero_date(pr.get("PayRunPeriodEndDate")),
                        "status": pr.get("PayRunStatus"),
                        "gross_wages": float(pr.get("Wages", 0) or 0),
                        "payg_withheld": float(pr.get("Tax", 0) or 0),
                        "super": float(pr.get("Super", 0) or 0),
                        "net_pay": float(pr.get("NetPay", 0) or 0),
                        "employee_count": len(pr.get("Payslips", [])),
                    }
                )

        # Sort by payment date
        pay_runs.sort(key=lambda x: x.get("payment_date", ""))
        return pay_runs

    except requests.RequestException as e:
        logger.warning("Failed to fetch pay runs: %s", e)
        return []


def _calculate_payroll_totals(pay_runs: list[dict]) -> dict[str, float]:
    """Calculate totals from pay runs."""
    total_gross = sum(pr.get("gross_wages", 0) for pr in pay_runs)
    total_payg = sum(pr.get("payg_withheld", 0) for pr in pay_runs)
    total_super = sum(pr.get("super", 0) for pr in pay_runs)
    total_net = sum(pr.get("net_pay", 0) for pr in pay_runs)
    total_employees = max(
        (pr.get("employee_count", 0) for pr in pay_runs),
        default=0,
    )

    return {
        "w1_gross_wages": round(total_gross, 2),
        "w2_payg_withheld": round(total_payg, 2),
        "super": round(total_super, 2),
        "net_pay": round(total_net, 2),
        "pay_run_count": len(pay_runs),
        "employee_count": total_employees,
    }


def _fetch_bas_data(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, Any]:
    """
    Fetch BAS data for the period.

    Note: Xero API doesn't directly expose BAS data.
    This attempts to get related report data or returns placeholder.
    """
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    # Try to fetch GST report which may have W1/W2 data
    try:
        resp = requests.get(
            f"{XERO_API_URL}/Reports/GST",
            headers=headers,
            params={"fromDate": from_date, "toDate": to_date},
            timeout=30,
        )

        if resp.status_code == 200:
            data = resp.json()
            # Parse GST report for wage-related labels
            return _parse_gst_report_for_wages(data)

    except requests.RequestException as e:
        logger.debug("Could not fetch GST report: %s", e)

    # Return placeholder - BAS data needs manual entry or isn't available
    return {
        "w1_gross_wages": None,
        "w2_payg_withheld": None,
        "source": "not_available",
        "message": "BAS data not available via API. Enter manually for comparison.",
    }


def _parse_gst_report_for_wages(report_data: dict) -> dict[str, Any]:
    """Parse GST report looking for wage-related fields."""
    # The GST report structure varies - this is a best-effort parse
    reports = report_data.get("Reports", [])
    if not reports:
        return {
            "w1_gross_wages": None,
            "w2_payg_withheld": None,
            "source": "gst_report",
            "message": "GST report available but wage data not found",
        }

    # Look for W1 and W2 labels in the report rows
    w1_value = None
    w2_value = None

    for report in reports:
        for row in report.get("Rows", []):
            cells = row.get("Cells", [])
            if len(cells) >= 2:
                label = str(cells[0].get("Value", "")).upper()
                if "W1" in label or "GROSS" in label:
                    try:
                        w1_value = float(cells[1].get("Value", 0))
                    except (ValueError, TypeError):
                        pass
                if "W2" in label or "WITHHELD" in label:
                    try:
                        w2_value = float(cells[1].get("Value", 0))
                    except (ValueError, TypeError):
                        pass

    return {
        "w1_gross_wages": w1_value,
        "w2_payg_withheld": w2_value,
        "source": "gst_report",
        "message": None if w2_value else "W2 data not found in GST report",
    }


def _calculate_variance(
    payroll: dict[str, Any],
    bas: dict[str, Any],
) -> dict[str, Any]:
    """Calculate variance between payroll and BAS."""
    payroll_w1 = payroll.get("w1_gross_wages", 0) or 0
    payroll_w2 = payroll.get("w2_payg_withheld", 0) or 0
    bas_w1 = bas.get("w1_gross_wages")
    bas_w2 = bas.get("w2_payg_withheld")

    variance = {
        "w1_variance": None,
        "w1_variance_pct": None,
        "w2_variance": None,
        "w2_variance_pct": None,
        "comparable": False,
    }

    if bas_w1 is not None:
        variance["w1_variance"] = round(payroll_w1 - bas_w1, 2)
        if bas_w1 != 0:
            variance["w1_variance_pct"] = round((payroll_w1 - bas_w1) / bas_w1 * 100, 2)
        variance["comparable"] = True

    if bas_w2 is not None:
        variance["w2_variance"] = round(payroll_w2 - bas_w2, 2)
        if bas_w2 != 0:
            variance["w2_variance_pct"] = round((payroll_w2 - bas_w2) / bas_w2 * 100, 2)
        variance["comparable"] = True

    return variance


def _determine_status(variance: dict[str, Any]) -> str:
    """Determine reconciliation status based on variance."""
    if not variance.get("comparable"):
        return "incomplete"

    w2_var = variance.get("w2_variance")
    if w2_var is None:
        return "incomplete"

    abs_var = abs(w2_var)
    if abs_var <= 1:  # Within $1
        return "ok"
    if abs_var <= 100:  # Within $100
        return "warning"
    return "error"


def _generate_warnings(variance: dict[str, Any], pay_runs: list[dict]) -> list[str]:
    """Generate warning messages."""
    warnings = []

    if not variance.get("comparable"):
        warnings.append(
            "BAS data not available for comparison. "
            "Enter W1/W2 values manually to complete reconciliation."
        )

    # Check for draft pay runs
    draft_runs = [pr for pr in pay_runs if pr.get("status") == "DRAFT"]
    if draft_runs:
        warnings.append(
            f"{len(draft_runs)} pay run(s) still in DRAFT status. "
            "Totals may change when posted."
        )

    # Check for variance
    w2_var = variance.get("w2_variance")
    if w2_var is not None and abs(w2_var) > 100:
        warnings.append(
            f"Significant W2 variance of ${abs(w2_var):,.2f}. "
            "Investigate discrepancy before lodging BAS."
        )

    return warnings


def _parse_xero_date(date_value: str | None) -> str | None:
    """Parse Xero date format /Date(timestamp)/ to YYYY-MM-DD."""
    if not date_value:
        return None

    if "/Date(" in str(date_value):
        try:
            ts = int(
                str(date_value).split("(")[1].split("+")[0].split("-")[0].split(")")[0]
            )
            dt = datetime.fromtimestamp(ts / 1000)
            return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            return None

    return str(date_value)


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export reconciliation data to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAYG-W Reconciliation"

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    # Title
    ws["A1"] = "PAYG-W Reconciliation Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    # Period
    period = data.get("period", {})
    ws["A2"] = f"Period: {period.get('from_date')} to {period.get('to_date')}"
    ws["A2"].font = Font(italic=True)

    # Summary section
    row = 4
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = header_font

    row += 1
    payroll = data.get("data", {}).get("payroll", {})
    bas = data.get("data", {}).get("bas", {})
    variance = data.get("data", {}).get("variance", {})

    summary_data = [
        ("", "Payroll", "BAS", "Variance"),
        (
            "W1 - Gross Wages",
            payroll.get("w1_gross_wages", 0),
            bas.get("w1_gross_wages") or "N/A",
            variance.get("w1_variance") or "N/A",
        ),
        (
            "W2 - PAYG Withheld",
            payroll.get("w2_payg_withheld", 0),
            bas.get("w2_payg_withheld") or "N/A",
            variance.get("w2_variance") or "N/A",
        ),
    ]

    for i, row_data in enumerate(summary_data):
        for j, value in enumerate(row_data):
            cell = ws.cell(row=row + i, column=j + 1, value=value)
            if i == 0:
                cell.font = header_font_white
                cell.fill = header_fill
            elif j > 0 and isinstance(value, int | float):
                cell.number_format = '"$"#,##0.00'

    row += len(summary_data) + 2

    # Pay runs detail
    ws[f"A{row}"] = "Pay Runs"
    ws[f"A{row}"].font = header_font
    row += 1

    pay_run_headers = [
        "Payment Date",
        "Period",
        "Status",
        "Gross Wages",
        "PAYG Withheld",
        "Super",
        "Net Pay",
        "Employees",
    ]

    for i, header in enumerate(pay_run_headers):
        cell = ws.cell(row=row, column=i + 1, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    row += 1
    pay_runs = data.get("data", {}).get("pay_runs", [])
    for pr in pay_runs:
        ws.cell(row=row, column=1, value=pr.get("payment_date"))
        period_str = f"{pr.get('period_start', '')} - {pr.get('period_end', '')}"
        ws.cell(row=row, column=2, value=period_str)
        ws.cell(row=row, column=3, value=pr.get("status"))
        ws.cell(
            row=row, column=4, value=pr.get("gross_wages", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=pr.get("payg_withheld", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=6, value=pr.get("super", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=7, value=pr.get("net_pay", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=8, value=pr.get("employee_count", 0))
        row += 1

    # Auto-adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
