"""
Fuel Tax Credits Calculator Service

Calculate FTC claims for eligible fuel purchases.
"""

import logging
from io import BytesIO
from typing import Any

import requests

from webapp.time_utils import utcnow_iso

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"

# FTC rates per litre (simplified - actual rates vary by period)
FTC_RATES = {
    "heavy_vehicle": 0.198,  # On-road heavy vehicles (>4.5 tonnes GVM)
    "light_vehicle": 0.198,  # Light vehicles for business
    "off_road": 0.488,  # Off-road/non-public roads
    "agriculture": 0.488,  # Agriculture, forestry, fishing
    "marine": 0.488,  # Marine vessels
    "rail": 0.488,  # Rail transport
}

# Common fuel supplier keywords
FUEL_SUPPLIERS = [
    "bp",
    "shell",
    "caltex",
    "ampol",
    "mobil",
    "7-eleven",
    "liberty",
    "united",
    "puma",
    "metro",
    "viva",
    "woolworths",
    "coles express",
]


def calculate_fuel_tax_credits(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
    fuel_type: str = "heavy_vehicle",
    manual_litres: float | None = None,
) -> dict[str, Any]:
    """
    Calculate fuel tax credits for the period.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)
        fuel_type: Type of fuel usage
        manual_litres: Optional manual litres entry

    Returns:
        Dict with FTC calculation
    """
    try:
        # Fetch fuel-related invoices
        fuel_invoices = _fetch_fuel_invoices(
            access_token, tenant_id, from_date, to_date
        )

        # Calculate totals
        fuel_spend = sum(inv.get("amount", 0) for inv in fuel_invoices)

        # Estimate litres if not manually provided
        if manual_litres is not None:
            estimated_litres = manual_litres
        else:
            # Rough estimate: assume average fuel price of $1.80/L
            estimated_litres = fuel_spend / 1.80 if fuel_spend > 0 else 0

        # Get FTC rate
        rate = FTC_RATES.get(fuel_type, FTC_RATES["heavy_vehicle"])

        # Calculate FTC
        ftc_claim = estimated_litres * rate

        return {
            "success": True,
            "data": {
                "fuel_invoices": fuel_invoices,
                "calculation": {
                    "fuel_spend": round(fuel_spend, 2),
                    "estimated_litres": round(estimated_litres, 2),
                    "litres_source": "manual" if manual_litres else "estimated",
                    "fuel_type": fuel_type,
                    "rate_per_litre": rate,
                    "ftc_claim": round(ftc_claim, 2),
                },
            },
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": utcnow_iso(),
        }

    except Exception as e:
        logger.exception("Error calculating fuel tax credits: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": utcnow_iso(),
        }


def _fetch_fuel_invoices(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> list[dict]:
    """Fetch invoices that appear to be fuel-related."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        # Fetch accounts payable invoices
        resp = requests.get(
            f"{XERO_API_URL}/Invoices",
            headers=headers,
            params={
                "where": f'Type=="ACCPAY" AND Date >= DateTime({from_date}) AND Date <= DateTime({to_date})',
                "order": "Date DESC",
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        fuel_invoices = []
        for inv in data.get("Invoices", []):
            contact_name = inv.get("Contact", {}).get("Name", "").lower()

            # Check if this appears to be a fuel supplier
            is_fuel = any(supplier in contact_name for supplier in FUEL_SUPPLIERS)

            # Also check line items for fuel-related descriptions
            if not is_fuel:
                for line in inv.get("LineItems", []):
                    desc = str(line.get("Description", "")).lower()
                    if any(
                        word in desc
                        for word in ["fuel", "diesel", "petrol", "unleaded", "lpg"]
                    ):
                        is_fuel = True
                        break

            if is_fuel:
                fuel_invoices.append(
                    {
                        "invoice_id": inv.get("InvoiceID"),
                        "invoice_number": inv.get("InvoiceNumber"),
                        "date": inv.get("Date", "")[:10] if inv.get("Date") else "",
                        "contact": inv.get("Contact", {}).get("Name", ""),
                        "amount": float(inv.get("Total", 0) or 0),
                        "gst": float(inv.get("TotalTax", 0) or 0),
                        "reference": inv.get("Reference", ""),
                    }
                )

        return fuel_invoices

    except requests.RequestException as e:
        logger.warning("Failed to fetch invoices: %s", e)
        return []


def get_ftc_rates() -> dict[str, dict]:
    """Return FTC rates for all fuel types."""
    return {
        "heavy_vehicle": {
            "rate": FTC_RATES["heavy_vehicle"],
            "description": "Heavy vehicles (>4.5t GVM) on public roads",
        },
        "light_vehicle": {
            "rate": FTC_RATES["light_vehicle"],
            "description": "Light vehicles for business use",
        },
        "off_road": {
            "rate": FTC_RATES["off_road"],
            "description": "Off-road/non-public road use",
        },
        "agriculture": {
            "rate": FTC_RATES["agriculture"],
            "description": "Agriculture, forestry, fishing",
        },
        "marine": {
            "rate": FTC_RATES["marine"],
            "description": "Marine vessels",
        },
        "rail": {
            "rate": FTC_RATES["rail"],
            "description": "Rail transport",
        },
    }


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export FTC calculation to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fuel Tax Credits"

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    result = data.get("data", {})
    period = data.get("period", {})

    # Title
    ws["A1"] = "Fuel Tax Credits Calculation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Period: {period.get('from_date')} to {period.get('to_date')}"
    ws["A2"].font = Font(italic=True)

    # Calculation summary
    row = 4
    ws[f"A{row}"] = "Calculation Summary"
    ws[f"A{row}"].font = Font(bold=True)

    calc = result.get("calculation", {})
    row += 1
    calc_items = [
        ("Total Fuel Spend", calc.get("fuel_spend", 0)),
        ("Estimated Litres", calc.get("estimated_litres", 0)),
        ("Litres Source", calc.get("litres_source", "")),
        ("Fuel Type", calc.get("fuel_type", "").replace("_", " ").title()),
        ("Rate per Litre", calc.get("rate_per_litre", 0)),
        ("FTC Claim", calc.get("ftc_claim", 0)),
    ]

    for label, value in calc_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if label in ("Total Fuel Spend", "FTC Claim"):
            cell.number_format = '"$"#,##0.00'
        elif label == "Rate per Litre":
            cell.number_format = '"$"0.000'
        row += 1

    row += 1

    # Fuel invoices
    ws[f"A{row}"] = "Fuel Invoices"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = ["Date", "Invoice #", "Supplier", "Amount", "GST"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for inv in result.get("fuel_invoices", []):
        ws.cell(row=row, column=1, value=inv.get("date", ""))
        ws.cell(row=row, column=2, value=inv.get("invoice_number", ""))
        ws.cell(row=row, column=3, value=inv.get("contact", ""))
        ws.cell(
            row=row, column=4, value=inv.get("amount", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=inv.get("gst", 0)
        ).number_format = '"$"#,##0.00'
        row += 1

    row += 2

    # Note
    ws.cell(
        row=row,
        column=1,
        value="Note: Litres are estimated. For accurate FTC claims, "
        "use actual litre data from fuel receipts.",
    )
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"].font = Font(italic=True, color="666666")

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
