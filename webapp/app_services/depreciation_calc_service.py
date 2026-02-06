"""
Depreciation Calculator Service

Calculate and review quarterly depreciation using Australian
diminishing value and prime cost methods.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import requests

logger = logging.getLogger(__name__)

XERO_API_URL = "https://api.xero.com/api.xro/2.0"

# Australian depreciation rates by asset class (effective life in years)
ASSET_EFFECTIVE_LIVES = {
    "computer_equipment": 4,
    "office_furniture": 10,
    "motor_vehicles": 8,
    "plant_equipment": 15,
    "buildings": 40,
    "low_value_pool": 1,  # Instant asset write-off threshold
    "other": 10,
}


def generate_depreciation_schedule(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, Any]:
    """
    Generate depreciation schedule for the period.

    Args:
        access_token: Xero OAuth access token
        tenant_id: Xero tenant ID
        from_date: Start date (YYYY-MM-DD)
        to_date: End date (YYYY-MM-DD)

    Returns:
        Dict with depreciation schedule and calculations
    """
    try:
        # Fetch fixed asset accounts
        asset_accounts = _fetch_fixed_asset_accounts(access_token, tenant_id)

        # Fetch balance sheet for asset values
        balances = _fetch_asset_balances(access_token, tenant_id, from_date, to_date)

        # Fetch depreciation journals
        depreciation_journals = _fetch_depreciation_journals(
            access_token, tenant_id, from_date, to_date
        )

        # Build depreciation schedule
        schedule = _build_depreciation_schedule(
            asset_accounts, balances, depreciation_journals, from_date, to_date
        )

        # Calculate totals
        totals = _calculate_totals(schedule)

        return {
            "success": True,
            "data": {
                "schedule": schedule,
                "totals": totals,
                "depreciation_journals": depreciation_journals,
            },
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.exception("Error generating depreciation schedule: %s", e)
        return {
            "success": False,
            "error": str(e),
            "data": None,
            "period": {"from_date": from_date, "to_date": to_date},
            "generated_at": datetime.utcnow().isoformat(),
        }


def _fetch_fixed_asset_accounts(
    access_token: str,
    tenant_id: str,
) -> list[dict]:
    """Fetch fixed asset accounts from Xero."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        resp = requests.get(
            f"{XERO_API_URL}/Accounts",
            headers=headers,
            params={"where": 'Class=="ASSET"'},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        accounts = []
        for acc in data.get("Accounts", []):
            acc_type = acc.get("Type", "")
            # Filter for fixed asset types
            if acc_type in ("FIXED", "NONCURRENT"):
                accounts.append(
                    {
                        "account_id": acc.get("AccountID"),
                        "code": acc.get("Code"),
                        "name": acc.get("Name"),
                        "type": acc_type,
                        "class": acc.get("Class"),
                        "asset_category": _determine_asset_category(
                            acc.get("Name", "")
                        ),
                    }
                )

        return accounts

    except requests.RequestException as e:
        logger.warning("Failed to fetch asset accounts: %s", e)
        return []


def _fetch_asset_balances(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> dict[str, dict]:
    """Fetch balance sheet to get asset values."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    balances = {}

    try:
        # Get opening balance (from_date)
        resp_open = requests.get(
            f"{XERO_API_URL}/Reports/BalanceSheet",
            headers=headers,
            params={"date": from_date},
            timeout=30,
        )
        if resp_open.status_code == 200:
            opening_balances = _parse_balance_sheet(resp_open.json())
            for code, value in opening_balances.items():
                balances[code] = {"opening": value, "closing": 0}

        # Get closing balance (to_date)
        resp_close = requests.get(
            f"{XERO_API_URL}/Reports/BalanceSheet",
            headers=headers,
            params={"date": to_date},
            timeout=30,
        )
        if resp_close.status_code == 200:
            closing_balances = _parse_balance_sheet(resp_close.json())
            for code, value in closing_balances.items():
                if code in balances:
                    balances[code]["closing"] = value
                else:
                    balances[code] = {"opening": 0, "closing": value}

    except requests.RequestException as e:
        logger.debug("Could not fetch balance sheet: %s", e)

    return balances


def _parse_balance_sheet(report_data: dict) -> dict[str, float]:
    """Parse balance sheet report for account balances."""
    balances: dict[str, float] = {}

    reports = report_data.get("Reports", [])
    if not reports:
        return balances

    def parse_rows(rows: list):
        for row in rows:
            row_type = row.get("RowType")

            if row_type == "Section":
                # Check if this is the Assets section
                title = row.get("Title", "")
                if "Asset" in title:
                    parse_rows(row.get("Rows", []))

            elif row_type == "Row":
                cells = row.get("Cells", [])
                if len(cells) >= 2:
                    # Extract account code from first cell
                    account_ref = cells[0].get("Value", "")
                    if " - " in account_ref:
                        account_code = account_ref.split(" - ")[0].strip()
                    else:
                        account_code = account_ref.strip()

                    try:
                        value = float(
                            str(cells[1].get("Value", 0))
                            .replace("$", "")
                            .replace(",", "")
                        )
                        if account_code:
                            balances[account_code] = value
                    except (ValueError, TypeError):
                        pass

    parse_rows(reports[0].get("Rows", []))
    return balances


def _fetch_depreciation_journals(
    access_token: str,
    tenant_id: str,
    from_date: str,
    to_date: str,
) -> list[dict]:
    """Fetch journal entries related to depreciation."""
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Xero-Tenant-Id": tenant_id,
        "Accept": "application/json",
    }

    try:
        # Search for journals with depreciation in reference/narration
        resp = requests.get(
            f"{XERO_API_URL}/Journals",
            headers=headers,
            params={"offset": "0", "paymentsOnly": "false"},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        from_dt = datetime.strptime(from_date, "%Y-%m-%d")
        to_dt = datetime.strptime(to_date, "%Y-%m-%d")

        journals = []
        for journal in data.get("Journals", []):
            journal_date_str = journal.get("JournalDate", "")
            if not journal_date_str:
                continue

            # Parse journal date
            try:
                journal_date = datetime.strptime(journal_date_str[:10], "%Y-%m-%d")
            except ValueError:
                continue

            if not (from_dt <= journal_date <= to_dt):
                continue

            # Check if this is a depreciation journal
            reference = str(journal.get("Reference", "")).lower()
            is_depreciation = (
                "depreciation" in reference
                or "depn" in reference
                or "dep'n" in reference
            )

            if is_depreciation:
                lines = []
                for line in journal.get("JournalLines", []):
                    lines.append(
                        {
                            "account_code": line.get("AccountCode"),
                            "account_name": line.get("AccountName"),
                            "debit": float(line.get("GrossAmount", 0) or 0)
                            if float(line.get("GrossAmount", 0) or 0) > 0
                            else 0,
                            "credit": abs(float(line.get("GrossAmount", 0) or 0))
                            if float(line.get("GrossAmount", 0) or 0) < 0
                            else 0,
                        }
                    )

                journals.append(
                    {
                        "journal_id": journal.get("JournalID"),
                        "journal_number": journal.get("JournalNumber"),
                        "date": journal_date_str[:10],
                        "reference": journal.get("Reference", ""),
                        "lines": lines,
                    }
                )

        return journals

    except requests.RequestException as e:
        logger.debug("Could not fetch journals: %s", e)
        return []


def _determine_asset_category(account_name: str) -> str:
    """Determine asset category from account name."""
    name_lower = account_name.lower()

    if any(word in name_lower for word in ["computer", "it", "software", "hardware"]):
        return "computer_equipment"
    if any(word in name_lower for word in ["furniture", "fitting", "office equipment"]):
        return "office_furniture"
    if any(word in name_lower for word in ["motor", "vehicle", "car", "truck"]):
        return "motor_vehicles"
    if any(word in name_lower for word in ["plant", "machinery", "equipment"]):
        return "plant_equipment"
    if any(word in name_lower for word in ["building", "property"]):
        return "buildings"

    return "other"


def _build_depreciation_schedule(
    accounts: list[dict],
    balances: dict[str, dict],
    journals: list[dict],
    from_date: str,
    to_date: str,
) -> list[dict]:
    """Build depreciation schedule from accounts and balances."""
    schedule = []

    # Calculate period months for pro-rata
    from_dt = datetime.strptime(from_date, "%Y-%m-%d")
    to_dt = datetime.strptime(to_date, "%Y-%m-%d")
    period_months = (to_dt.year - from_dt.year) * 12 + (to_dt.month - from_dt.month) + 1

    for account in accounts:
        code = account.get("code", "")
        balance_data = balances.get(code, {})
        opening = balance_data.get("opening", 0)
        closing = balance_data.get("closing", 0)

        # Skip if no balance
        if opening == 0 and closing == 0:
            continue

        # Skip accumulated depreciation accounts (negative balances)
        if opening < 0 or closing < 0:
            continue

        category = account.get("asset_category", "other")
        effective_life = ASSET_EFFECTIVE_LIVES.get(category, 10)

        # Calculate expected depreciation (diminishing value method)
        # Rate = (200% / effective life) for diminishing value
        annual_rate = 2.0 / effective_life
        expected_annual = opening * annual_rate
        expected_quarterly = expected_annual * (period_months / 12)

        # Find actual depreciation from journals
        actual_depreciation = _find_actual_depreciation(journals, code)

        # Calculate variance
        variance = actual_depreciation - expected_quarterly
        variance_pct = (
            (variance / expected_quarterly * 100) if expected_quarterly else 0
        )

        # Determine status
        status = "ok"
        if abs(variance_pct) > 25:
            status = "error"
        elif abs(variance_pct) > 10:
            status = "warning"

        schedule.append(
            {
                "account_code": code,
                "account_name": account.get("name", ""),
                "category": category,
                "effective_life": effective_life,
                "depreciation_rate": round(annual_rate * 100, 1),
                "opening_value": round(opening, 2),
                "additions": round(max(closing - opening + actual_depreciation, 0), 2),
                "depreciation_expected": round(expected_quarterly, 2),
                "depreciation_actual": round(actual_depreciation, 2),
                "variance": round(variance, 2),
                "variance_pct": round(variance_pct, 1),
                "closing_value": round(closing, 2),
                "status": status,
            }
        )

    return schedule


def _find_actual_depreciation(journals: list[dict], account_code: str) -> float:
    """Find actual depreciation amount from journals for an account."""
    total = 0.0

    for journal in journals:
        for line in journal.get("lines", []):
            line_code = line.get("account_code", "")
            # Match the asset account or its accumulated depreciation account
            if line_code == account_code or "accum" in line_code.lower():
                # Credit to asset account is depreciation
                total += line.get("credit", 0)

    return total


def _calculate_totals(schedule: list[dict]) -> dict[str, Any]:
    """Calculate totals from depreciation schedule."""
    total_opening = sum(item.get("opening_value", 0) for item in schedule)
    total_additions = sum(item.get("additions", 0) for item in schedule)
    total_expected = sum(item.get("depreciation_expected", 0) for item in schedule)
    total_actual = sum(item.get("depreciation_actual", 0) for item in schedule)
    total_closing = sum(item.get("closing_value", 0) for item in schedule)

    return {
        "total_opening": round(total_opening, 2),
        "total_additions": round(total_additions, 2),
        "total_depreciation_expected": round(total_expected, 2),
        "total_depreciation_actual": round(total_actual, 2),
        "total_variance": round(total_actual - total_expected, 2),
        "total_closing": round(total_closing, 2),
        "asset_count": len(schedule),
        "assets_ok": sum(1 for item in schedule if item.get("status") == "ok"),
        "assets_warning": sum(
            1 for item in schedule if item.get("status") == "warning"
        ),
        "assets_error": sum(1 for item in schedule if item.get("status") == "error"),
    }


def calculate_depreciation(
    asset_value: float,
    effective_life: float,
    method: str = "diminishing",
    period_months: int = 3,
) -> dict[str, float]:
    """
    Calculate depreciation using Australian tax methods.

    Args:
        asset_value: Current written down value
        effective_life: Effective life in years
        method: "diminishing" or "prime_cost"
        period_months: Period length in months

    Returns:
        Dict with annual and period depreciation amounts
    """
    if method == "diminishing":
        # Diminishing value: 200% / effective life
        annual_rate = 2.0 / effective_life
    else:
        # Prime cost: 100% / effective life
        annual_rate = 1.0 / effective_life

    annual_depreciation = asset_value * annual_rate
    period_depreciation = annual_depreciation * (period_months / 12)

    return {
        "annual_rate": round(annual_rate * 100, 2),
        "annual_depreciation": round(annual_depreciation, 2),
        "period_depreciation": round(period_depreciation, 2),
    }


def export_to_excel(data: dict[str, Any]) -> BytesIO:
    """Export depreciation schedule to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise ImportError("openpyxl required for Excel export") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depreciation Schedule"

    # Styles
    header_fill = PatternFill(
        start_color="0066CC", end_color="0066CC", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    warning_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )
    error_fill = PatternFill(
        start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
    )

    result = data.get("data", {})
    period = data.get("period", {})

    # Title
    ws["A1"] = "Depreciation Schedule"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Period: {period.get('from_date')} to {period.get('to_date')}"
    ws["A2"].font = Font(italic=True)

    # Summary
    row = 4
    ws[f"A{row}"] = "Summary"
    ws[f"A{row}"].font = Font(bold=True)

    totals = result.get("totals", {})
    row += 1
    summary_items = [
        ("Total Opening Value", totals.get("total_opening", 0)),
        ("Total Additions", totals.get("total_additions", 0)),
        ("Expected Depreciation", totals.get("total_depreciation_expected", 0)),
        ("Actual Depreciation", totals.get("total_depreciation_actual", 0)),
        ("Variance", totals.get("total_variance", 0)),
        ("Total Closing Value", totals.get("total_closing", 0)),
    ]

    for label, value in summary_items:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = '"$"#,##0.00'
        row += 1

    row += 1

    # Schedule table
    ws[f"A{row}"] = "Asset Schedule"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    headers = [
        "Account",
        "Category",
        "Rate %",
        "Opening",
        "Additions",
        "Expected Dep.",
        "Actual Dep.",
        "Variance",
        "Closing",
        "Status",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row += 1
    for item in result.get("schedule", []):
        status = item.get("status", "unknown")
        fill = (
            ok_fill
            if status == "ok"
            else warning_fill
            if status == "warning"
            else error_fill
        )

        ws.cell(row=row, column=1, value=item.get("account_name", ""))
        ws.cell(row=row, column=2, value=item.get("category", ""))
        ws.cell(row=row, column=3, value=item.get("depreciation_rate", 0))
        ws.cell(
            row=row, column=4, value=item.get("opening_value", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=5, value=item.get("additions", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=6, value=item.get("depreciation_expected", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=7, value=item.get("depreciation_actual", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=8, value=item.get("variance", 0)
        ).number_format = '"$"#,##0.00'
        ws.cell(
            row=row, column=9, value=item.get("closing_value", 0)
        ).number_format = '"$"#,##0.00'
        cell = ws.cell(row=row, column=10, value=status.upper())
        cell.fill = fill

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
