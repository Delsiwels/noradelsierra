"""
Payroll Review Blueprint

Page route and REST API endpoints for payroll review features:
- Pay run comparison (DRAFT vs POSTED)
- Leave flags with balance warnings
- Employee Excel upload and creation

Endpoints:
- GET  /payroll-review              - Render main page
- GET  /payroll-review/api/pay-runs - Get DRAFT and recent POSTED pay runs
- GET  /payroll-review/api/compare  - Compare two pay runs
- GET  /payroll-review/api/leave-flags - Get leave data for a pay run
- POST /payroll-review/api/upload-employees - Upload and parse Excel file
- POST /payroll-review/api/create-employees - Create parsed employees in Xero
- GET  /payroll-review/api/employee-template - Download blank Excel template
"""

import logging
import os
from functools import wraps

from flask import (
    Blueprint,
    current_app,
    jsonify,
    render_template,
    request,
    send_file,
    session,
)

from webapp.app_services.payroll_review_service import (
    build_leave_flags_response,
    compare_pay_runs,
    create_employees_in_xero,
    get_draft_pay_runs,
    get_employee_leave_balances,
    get_pay_run_with_payslips,
    get_recent_posted_pay_run,
    parse_employee_excel,
    validate_employee_data,
)

logger = logging.getLogger(__name__)

payroll_review_bp = Blueprint("payroll_review", __name__, url_prefix="/payroll-review")


def _get_current_user():
    """Get current authenticated user."""
    if current_app.config.get("TESTING"):
        return None
    try:
        from flask_login import current_user

        if current_user.is_authenticated:
            return current_user
    except (ImportError, AttributeError):
        pass
    return None


def _login_required(f):
    """Require login decorator. Bypassed in testing mode."""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_app.config.get("TESTING"):
            return f(*args, **kwargs)
        try:
            from flask_login import current_user

            if not current_user.is_authenticated:
                return jsonify({"error": "Authentication required"}), 401
        except (ImportError, AttributeError):
            pass
        return f(*args, **kwargs)

    return decorated_function


def _get_xero_credentials() -> tuple[str | None, str | None]:
    """Get Xero access token and tenant ID from session."""
    # Try new session structure first
    conn = session.get("xero_connection", {})
    access_token = conn.get("access_token") or session.get("xero_access_token")
    tenant_id = conn.get("tenant_id") or session.get("xero_tenant_id")
    return access_token, tenant_id


# =============================================================================
# Page Route
# =============================================================================


@payroll_review_bp.route("/")
@_login_required
def index():
    """Render the payroll review page."""
    return render_template("payroll_review.html")


# =============================================================================
# API Routes
# =============================================================================


@payroll_review_bp.route("/api/pay-runs", methods=["GET"])
@_login_required
def api_pay_runs():
    """
    Get DRAFT and recent POSTED pay runs.

    Response:
        - draft_pay_runs: list of DRAFT pay runs
        - recent_posted: most recent POSTED pay run or null
    """
    access_token, tenant_id = _get_xero_credentials()

    if not access_token or not tenant_id:
        return jsonify({"error": "Xero not connected"}), 400

    try:
        draft_pay_runs = get_draft_pay_runs(access_token, tenant_id)
        recent_posted = get_recent_posted_pay_run(access_token, tenant_id)

        return jsonify(
            {
                "success": True,
                "draft_pay_runs": draft_pay_runs,
                "recent_posted": recent_posted,
            }
        )
    except Exception as e:
        logger.exception("Error fetching pay runs: %s", e)
        return jsonify({"error": "Failed to fetch pay runs"}), 500


@payroll_review_bp.route("/api/compare", methods=["GET"])
@_login_required
def api_compare():
    """
    Compare a DRAFT pay run with a POSTED pay run.

    Query params:
        - draft_id: DRAFT pay run ID
        - posted_id: POSTED pay run ID (optional, uses most recent if not provided)

    Response:
        - draft: summary of draft pay run
        - posted: summary of posted pay run
        - comparison: list of employee variance rows
    """
    access_token, tenant_id = _get_xero_credentials()

    if not access_token or not tenant_id:
        return jsonify({"error": "Xero not connected"}), 400

    draft_id = request.args.get("draft_id")
    posted_id = request.args.get("posted_id")

    if not draft_id:
        return jsonify({"error": "draft_id is required"}), 400

    try:
        # Fetch draft pay run with payslips
        draft_detail = get_pay_run_with_payslips(access_token, tenant_id, draft_id)
        if not draft_detail:
            return jsonify({"error": "Draft pay run not found"}), 404

        # Fetch posted pay run
        posted_detail = None
        if posted_id:
            posted_detail = get_pay_run_with_payslips(
                access_token, tenant_id, posted_id
            )
        else:
            recent = get_recent_posted_pay_run(access_token, tenant_id)
            if recent:
                posted_detail = get_pay_run_with_payslips(
                    access_token, tenant_id, recent["pay_run_id"]
                )

        if not posted_detail:
            return jsonify(
                {
                    "success": True,
                    "draft": _format_pay_run_summary(draft_detail),
                    "posted": None,
                    "comparison": [],
                    "message": "No posted pay run available for comparison",
                }
            )

        # Run comparison
        comparison = compare_pay_runs(draft_detail, posted_detail)

        return jsonify(
            {
                "success": True,
                "draft": _format_pay_run_summary(draft_detail),
                "posted": _format_pay_run_summary(posted_detail),
                "comparison": comparison,
            }
        )

    except Exception as e:
        logger.exception("Error comparing pay runs: %s", e)
        return jsonify({"error": "Failed to compare pay runs"}), 500


@payroll_review_bp.route("/api/leave-flags", methods=["GET"])
@_login_required
def api_leave_flags():
    """
    Get leave data for a pay run with balance warnings.

    Query params:
        - pay_run_id: Pay run ID to analyze

    Response:
        - pay_run_id: the analyzed pay run
        - employees_with_leave: list of employees with leave taken
    """
    access_token, tenant_id = _get_xero_credentials()

    if not access_token or not tenant_id:
        return jsonify({"error": "Xero not connected"}), 400

    pay_run_id = request.args.get("pay_run_id")
    if not pay_run_id:
        return jsonify({"error": "pay_run_id is required"}), 400

    try:
        # Fetch pay run with payslips
        pay_run = get_pay_run_with_payslips(access_token, tenant_id, pay_run_id)
        if not pay_run:
            return jsonify({"error": "Pay run not found"}), 404

        payslips = pay_run.get("payslips", [])

        # Get unique employee IDs that have leave
        employee_ids = set()
        for ps in payslips:
            if ps.get("LeaveEarningsLines"):
                emp_id = ps.get("EmployeeID")
                if emp_id:
                    employee_ids.add(emp_id)

        # Fetch leave balances
        leave_balances = {}
        if employee_ids:
            leave_balances = get_employee_leave_balances(
                access_token, tenant_id, list(employee_ids)
            )

        # Build response
        employees_with_leave = build_leave_flags_response(payslips, leave_balances)

        return jsonify(
            {
                "success": True,
                "pay_run_id": pay_run_id,
                "employees_with_leave": employees_with_leave,
            }
        )

    except Exception as e:
        logger.exception("Error fetching leave flags: %s", e)
        return jsonify({"error": "Failed to fetch leave data"}), 500


@payroll_review_bp.route("/api/upload-employees", methods=["POST"])
@_login_required
def api_upload_employees():
    """
    Upload and parse an Excel file containing employee data.

    Request: multipart/form-data with 'file' field

    Response:
        - parsed_count: number of rows parsed
        - valid_count: number of valid rows
        - employees: list of parsed employee data with validation status
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    # Validate file extension
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "File must be an Excel file (.xlsx or .xls)"}), 400

    try:
        file_data = file.read()

        # Limit file size (5MB)
        if len(file_data) > 5 * 1024 * 1024:
            return jsonify({"error": "File size exceeds 5MB limit"}), 400

        result = parse_employee_excel(file_data)

        if not result.get("success", False):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": result.get("error", "Failed to parse file"),
                    }
                ),
                400,
            )

        return jsonify(result)

    except Exception as e:
        logger.exception("Error uploading employee file: %s", e)
        return jsonify({"error": "Failed to process uploaded file"}), 500


@payroll_review_bp.route("/api/create-employees", methods=["POST"])
@_login_required
def api_create_employees():
    """
    Create employees in Xero from previously parsed data.

    Request (JSON):
        - employees: list of employee dicts to create

    Response:
        - total: total employees processed
        - created: number successfully created
        - failed: number that failed
        - results: per-employee results
    """
    access_token, tenant_id = _get_xero_credentials()

    if not access_token or not tenant_id:
        return jsonify({"error": "Xero not connected"}), 400

    data = request.get_json(silent=True)
    if not data or "employees" not in data:
        return jsonify({"error": "employees list is required"}), 400

    employees = data["employees"]
    if not isinstance(employees, list):
        return jsonify({"error": "employees must be a list"}), 400

    if len(employees) > 50:
        return jsonify({"error": "Maximum 50 employees per batch"}), 400

    try:
        # Re-validate before creation
        validated = validate_employee_data(employees)

        # Create in Xero
        result = create_employees_in_xero(access_token, tenant_id, validated)

        return jsonify(result)

    except Exception as e:
        logger.exception("Error creating employees: %s", e)
        return jsonify({"error": "Failed to create employees"}), 500


@payroll_review_bp.route("/api/employee-template", methods=["GET"])
@_login_required
def api_employee_template():
    """
    Download the blank employee Excel template.

    Returns the template file as a download.
    """
    template_path = os.path.join(
        current_app.root_path, "static", "templates", "employee_template.xlsx"
    )

    if not os.path.exists(template_path):
        # Generate template on-the-fly if it doesn't exist
        try:
            template_data = _generate_employee_template()
            return send_file(
                template_data,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="employee_template.xlsx",
            )
        except Exception as e:
            logger.exception("Error generating template: %s", e)
            return jsonify({"error": "Template not available"}), 500

    return send_file(
        template_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="employee_template.xlsx",
    )


# =============================================================================
# Helper Functions
# =============================================================================


def _format_pay_run_summary(pay_run: dict) -> dict:
    """Format a pay run for summary display."""
    start = pay_run.get("pay_run_period_start_date", "")
    end = pay_run.get("pay_run_period_end_date", "")

    period = ""
    if start and end:
        try:
            from datetime import datetime

            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d")
            period = f"{start_dt.strftime('%d %b')} - {end_dt.strftime('%d %b %Y')}"
        except ValueError:
            period = f"{start} to {end}"

    return {
        "pay_run_id": pay_run.get("pay_run_id"),
        "payment_date": pay_run.get("payment_date"),
        "period": period,
        "status": pay_run.get("status"),
        "total_gross": pay_run.get("wages", 0),
        "total_super": pay_run.get("super", 0),
        "total_tax": pay_run.get("tax", 0),
        "total_net": pay_run.get("net_pay", 0),
        "employee_count": len(pay_run.get("payslips", [])),
    }


def _generate_employee_template():
    """Generate an Excel template file on-the-fly."""
    from io import BytesIO

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise ImportError("openpyxl required for template generation") from err

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Headers
    headers = [
        "First Name",
        "Last Name",
        "Date of Birth",
        "Email",
        "Phone",
        "Address Line 1",
        "City",
        "State",
        "Postcode",
        "Start Date",
        "Job Title",
        "TFN",
        "Bank BSB",
        "Bank Account Number",
        "Bank Account Name",
        "Super Fund USI",
        "Super Member Number",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Set column widths
    column_widths = {
        "First Name": 15,
        "Last Name": 15,
        "Date of Birth": 14,
        "Email": 25,
        "Phone": 15,
        "Address Line 1": 25,
        "City": 15,
        "State": 8,
        "Postcode": 10,
        "Start Date": 14,
        "Job Title": 20,
        "TFN": 12,
        "Bank BSB": 10,
        "Bank Account Number": 18,
        "Bank Account Name": 20,
        "Super Fund USI": 20,
        "Super Member Number": 18,
    }

    for col, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = column_widths.get(
            header, 15
        )

    # Add data validation for State column
    from openpyxl.worksheet.datavalidation import DataValidation

    state_col = headers.index("State") + 1
    state_validation = DataValidation(
        type="list",
        formula1='"NSW,VIC,QLD,SA,WA,TAS,NT,ACT"',
        allow_blank=True,
    )
    state_validation.error = "Please select a valid Australian state"
    state_validation.errorTitle = "Invalid State"
    ws.add_data_validation(state_validation)
    state_validation.add(
        f"{get_column_letter(state_col)}2:{get_column_letter(state_col)}100"
    )

    # Add example row (commented out data)
    example_data = [
        "John",
        "Smith",
        "15/03/1990",
        "john.smith@example.com",
        "0412345678",
        "123 Main Street",
        "Sydney",
        "NSW",
        "2000",
        "01/02/2026",
        "Accountant",
        "123456789",
        "062000",
        "12345678",
        "J Smith",
        "STA0100AU",
        "12345",
    ]

    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = openpyxl.styles.Font(color="808080", italic=True)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
